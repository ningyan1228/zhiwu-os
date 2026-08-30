"""Import the human-verified CPPH-2A strict list without rewriting its evidence."""
from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Awaitable, Callable
from urllib.parse import quote, urlparse

from openpyxl import load_workbook

STRICT_SHEET = "严格客户名单"
STRICT_TASK_NAME = "CPPH-2A全球严格"
STRICT_DATA_SOURCE = "manual_verified_excel"
STRICT_VERIFICATION_STATUS = "strict_verified_import"
CONFIRMATION_NOTE = "公开证据仅证明潜在功能匹配，不代表企业已采购或已批准使用CPPH-2A/CPP/CPO。"
PENDING_REASON = "自动抓取记录，未通过人工严格核验导入标准；需重新核验企业主体、联系方式与官方产品证据。"

HEADERS = {
    "ID": "sourceRecordId", "匹配级别": "matchLevel", "国家/地区": "countryRegion",
    "公司名称": "companyName", "企业类型": "companyType", "官网": "officialWebsite",
    "公开联系人/部门": "publicContactOrDepartment", "公开邮箱": "publicEmail",
    "公开电话": "publicPhone", "公开地址/范围": "addressOrBusinessScope",
    "PP/PE/TPO相关公开证据": "productApplicationEvidence", "应用范围": "applicationScope",
    "与H-2A的潜在匹配点": "potentialFit", "推荐首联部门": "recommendedContactDepartment",
    "首轮需确认问题": "firstContactQuestions", "真实性核验结论": "verificationConclusion",
    "来源链接": "sourceUrls",
}
REQUIRED_HEADERS = tuple(HEADERS)


def text(value: Any) -> str:
    return str(value or "").strip()


def url_domain(value: str) -> str:
    return (urlparse(value).hostname or "").lower().removeprefix("www.")


def split_source_urls(value: str) -> list[str]:
    """Keep the Excel's URLs in order, removing only blank duplicate values."""
    urls: list[str] = []
    for candidate in re.split(r"[\r\n]+", value):
        candidate = candidate.strip()
        if candidate and candidate not in urls:
            urls.append(candidate)
    return urls


def grade(match_level: str) -> str:
    if match_level.startswith("A"):
        return "A"
    if match_level.startswith("B"):
        return "B"
    raise ValueError(f"不支持的匹配级别：{match_level}")


def parse_cpph_strict_workbook(content: bytes) -> list[dict[str, str | list[str]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    if STRICT_SHEET not in workbook.sheetnames:
        raise ValueError(f"Excel 缺少工作表“{STRICT_SHEET}”")
    sheet = workbook[STRICT_SHEET]
    actual_headers = [text(sheet.cell(3, column).value) for column in range(1, 18)]
    if actual_headers != list(REQUIRED_HEADERS):
        raise ValueError("严格客户名单的 A3:Q3 字段与约定映射不一致")
    records: list[dict[str, str | list[str]]] = []
    for row_index in range(4, 45):
        values = [text(sheet.cell(row_index, column).value) for column in range(1, 18)]
        if not any(values):
            continue
        if not all(values):
            missing = ", ".join(REQUIRED_HEADERS[index] for index, value in enumerate(values) if not value)
            raise ValueError(f"第 {row_index} 行缺少严格核验字段：{missing}")
        record = {HEADERS[REQUIRED_HEADERS[index]]: values[index] for index in range(17)}
        record["sourceUrls"] = split_source_urls(values[16])
        if not record["sourceUrls"]:
            raise ValueError(f"第 {row_index} 行缺少可点击来源链接")
        if grade(str(record["matchLevel"])) not in {"A", "B"}:
            raise ValueError(f"第 {row_index} 行匹配级别无效")
        records.append(record)
    if len(records) != 41:
        raise ValueError(f"严格客户名单必须包含 A4:Q44 的 41 家企业，当前读取到 {len(records)} 家")
    if len({str(record['sourceRecordId']) for record in records}) != len(records):
        raise ValueError("Excel 的 ID 存在重复，不能幂等导入")
    return records


def lead_payload(record: dict[str, str | list[str]], task_id: str, now: str) -> dict[str, Any]:
    source_urls = record["sourceUrls"]
    assert isinstance(source_urls, list)
    official_website = str(record["officialWebsite"])
    contact_or_department = str(record["publicContactOrDepartment"])
    product_evidence = str(record["productApplicationEvidence"])
    source_url = official_website or source_urls[0]
    return {
        "task_id": task_id,
        "source_record_id": str(record["sourceRecordId"]),
        "match_level": str(record["matchLevel"]),
        "company_name": str(record["companyName"]),
        "country": str(record["countryRegion"]),
        "company_type": str(record["companyType"]),
        "website": official_website,
        "website_domain": url_domain(official_website),
        "official_website": official_website,
        "official_homepage_url": official_website,
        "source_url": source_url,
        "source_type": "其他公开网页",
        "first_discovery_source_url": source_urls[0],
        "official_validation_source_url": official_website,
        "company_source_url": official_website,
        "public_contact_or_department": contact_or_department,
        "contact_department": contact_or_department,
        "public_business_email": str(record["publicEmail"]),
        "public_business_phone": str(record["publicPhone"]),
        "official_address": str(record["addressOrBusinessScope"]),
        "business_scope": str(record["addressOrBusinessScope"]),
        "product_application_evidence": product_evidence,
        "product_evidence_summary": product_evidence,
        "product_evidence_type": "官网/官方PDF公开证据（人工严格核验）",
        "application_scope": str(record["applicationScope"]),
        "discovered_application_keywords": [str(record["applicationScope"])],
        "potential_fit": str(record["potentialFit"]),
        "possible_need": str(record["potentialFit"]),
        "recommended_contact_department": str(record["recommendedContactDepartment"]),
        "first_contact_questions": str(record["firstContactQuestions"]),
        "verification_conclusion": str(record["verificationConclusion"]),
        "source_urls": source_urls,
        "contact_source_url": next((url for url in source_urls if "contact" in url.lower()), source_urls[0]),
        "email_source_url": next((url for url in source_urls if "contact" in url.lower()), source_urls[0]),
        "phone_source_url": next((url for url in source_urls if "contact" in url.lower()), source_urls[0]),
        "product_evidence_url": next((url for url in source_urls if url != official_website), source_urls[0]),
        "matching_grade": grade(str(record["matchLevel"])),
        "verification_bucket": "严格客户名单",
        "verification_status": STRICT_VERIFICATION_STATUS,
        "data_source": STRICT_DATA_SOURCE,
        "imported_at": now,
        "verified_at": now,
        "needs_human_confirmation": True,
        "confirmation_note": CONFIRMATION_NOTE,
        "missing_requirements": [],
        "score_reasons": ["人工严格核验 Excel 导入；官网、联系人/部门、邮箱、电话及产品/应用证据均保留原文与来源链接。"],
        "match_score": 100,
        "suspected_duplicate": False,
        "robots_status": "manual_verified",
        "robots_reason": "由人工严格核验 Excel 导入；未由自动抓取推断。",
        "status": "保留",
        "updated_at": now,
    }


Request = Callable[[str, str, dict[str, Any] | None], Awaitable[Any]]


async def import_cpph_strict_records(request: Request, content: bytes) -> dict[str, int]:
    """Upsert the supplied strict workbook for the authenticated user.

    De-duplication follows the requested order: official domain, company name,
    public email, then address.  sourceRecordId makes repeat imports idempotent.
    """
    records = parse_cpph_strict_workbook(content)
    now = datetime.now(timezone.utc).isoformat()
    task_rows = await request(
        f"lead_search_tasks?task_name=eq.{quote(STRICT_TASK_NAME, safe='')}&deleted_at=is.null&select=*&limit=1", "GET", None,
    )
    if task_rows:
        task = task_rows[0]
    else:
        task = (await request("lead_search_tasks", "POST", {
            "task_name": STRICT_TASK_NAME,
            "product_keywords": ["CPPH-2A", "CPP", "CPO", "chlorinated polyolefin adhesion promoter"],
            "application_keywords": ["PP", "PE", "TPO", "ink", "coating"],
            "target_countries": [], "excluded_countries": [],
            "target_company_types": ["制造商", "配方商", "加工厂", "品牌方"],
            "source_urls": [], "search_language": "English", "max_results": 41,
            "daily_enabled": False, "daily_run_time": "08:30", "status": "暂停",
        }))[0]
    known = await request("customer_leads?select=*&limit=500", "GET", None)
    inserted = updated = 0
    for record in records:
        payload = lead_payload(record, task["id"], now)
        domain = payload["website_domain"]
        name = str(payload["company_name"]).casefold()
        email = str(payload["public_business_email"]).casefold()
        address = str(payload["official_address"]).casefold()
        source_id = str(payload["source_record_id"])
        existing = next((item for item in known if item.get("task_id") == task["id"] and item.get("data_source") == STRICT_DATA_SOURCE and str(item.get("source_record_id") or "") == source_id), None)
        if not existing and domain:
            existing = next((item for item in known if url_domain(str(item.get("official_website") or item.get("official_homepage_url") or item.get("website") or "")) == domain), None)
        if not existing:
            existing = next((item for item in known if str(item.get("company_name") or "").casefold() == name), None)
        if not existing and email:
            existing = next((item for item in known if str(item.get("public_business_email") or "").casefold() == email), None)
        if not existing and address:
            existing = next((item for item in known if str(item.get("official_address") or item.get("business_scope") or "").casefold() == address), None)
        if existing:
            await request(f"customer_leads?id=eq.{existing['id']}", "PATCH", payload)
            existing.update(payload)
            updated += 1
        else:
            created = (await request("customer_leads", "POST", payload))[0]
            known.append(created)
            inserted += 1
    return {"total": len(records), "inserted": inserted, "updated": updated, "a_count": sum(grade(str(item["matchLevel"])) == "A" for item in records), "b_count": sum(grade(str(item["matchLevel"])) == "B" for item in records)}
