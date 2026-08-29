"""Thin API gateway: browser credentials are verified with Supabase before data is proxied."""
from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
from typing import Any, Literal
from urllib.parse import quote, urlparse

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font
from fastapi import BackgroundTasks, FastAPI, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel, Field
from pydantic_settings import BaseSettings, SettingsConfigDict

class Settings(BaseSettings):
    supabase_url: str
    supabase_anon_key: str
    supabase_service_role_key: str
    allowed_origins: str = "http://localhost:5173"
    public_app_url: str = "https://ningyan1228.github.io/zhiwu-os/"
    mail_host: str | None = None
    mail_port: int = 993
    mail_username: str | None = None
    mail_password: str | None = None
    mail_folder: str = "INBOX"
    mail_owner_user_id: str | None = None
    mail_internal_addresses: str = ""
    mail_internal_domains: str = ""
    mail_sync_interval_seconds: int = 600
    mail_sync_max_messages: int = 100
    lead_discovery_user_agent: str = "ZhiwuOSLeadDiscovery/1.0 (+https://work.101921.xyz)"
    lead_discovery_delay_seconds: float = 1.0
    # Optional official search provider token. It is read only on the server and
    # must never be returned by an API endpoint or committed to the repository.
    brave_search_api_key: str | None = None
    model_config = SettingsConfigDict(env_file=".env", extra="ignore")

@lru_cache
def settings() -> Settings: return Settings()

app = FastAPI(title="Zhiwu OS API", version="0.1.0")
app.add_middleware(CORSMiddleware, allow_origins=settings().allowed_origins.split(","), allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

class CustomerIn(BaseModel):
    company_name: str
    country: str
    contact_person: str
    email: str
    whatsapp: str | None = None
    product_interest: str | None = None
    customer_stage: str = "New"
    priority: str = "MEDIUM"
    application: str | None = None
    status_label: str | None = None
    status_tone: str | None = None
    next_followup_date: str | None = None
    notes: str | None = None
    website: str | None = None
    wechat: str | None = None
    industry: str | None = None
    customer_summary: str | None = None
    customer_background: str | None = None
    customer_need: str | None = None
    important_notes: str | None = None
    customer_value: int | None = Field(default=None, ge=1, le=5)
    customer_tags: list[str] | None = None
    next_action: list[str] | None = None

class LoginIn(BaseModel):
    email: str
    password: str

class PasswordIn(BaseModel):
    password: str = Field(min_length=8, max_length=128)

class RecoveryIn(BaseModel):
    email: str

class ProductIn(BaseModel):
    product_name: str
    product_code: str
    category: str | None = None
    application: str | None = None
    description: str | None = None
    notes: str | None = None

class FollowupIn(BaseModel):
    customer_id: str
    date: str
    content: str
    next_action: str | None = None
    status: str = "Open"

class ProjectIn(BaseModel):
    customer_id: str
    project_name: str
    product_id: str | None = None
    application: str | None = None
    stage: str = "New Inquiry"
    notes: str | None = None

class ProjectUpdateIn(BaseModel):
    project_name: str
    product_id: str | None = None
    application: str | None = None
    stage: str = "New Inquiry"
    notes: str | None = None

class ProductCustomerRelationIn(BaseModel):
    customer_id: str

class QuoteIn(BaseModel):
    customer_id: str
    product_id: str | None = None
    quantity: str
    amount: float | None = None
    currency: str = "USD"
    trade_term: str | None = None
    status: str = "Draft"

SupplierStatus = Literal["待联系", "已询价", "等 TDS", "等报价", "等样品", "技术评估", "已合作", "暂停", "淘汰"]

class SupplierIn(BaseModel):
    company_name: str = Field(min_length=1, max_length=300)
    english_name: str | None = None
    country: str = "China"
    province: str | None = None
    city: str | None = None
    address: str | None = None
    website: str | None = None
    supplier_type: Literal["工厂", "贸易商", "待确认"] = "待确认"
    export_status: Literal["可出口", "不可出口", "待确认"] = "待确认"
    main_phone: str | None = None
    main_email: str | None = None
    wechat: str | None = None
    product_keywords: list[str] = []
    product_categories: list[str] = []
    supplier_tags: list[str] = []
    current_status: SupplierStatus = "待联系"
    last_contact_date: str | None = None
    next_action: str | None = None
    next_followup_date: str | None = None
    notes: str | None = None
    risk_notes: str | None = None

class SupplierUpdateIn(SupplierIn):
    pass

class SupplierContactIn(BaseModel):
    supplier_id: str
    name: str = Field(min_length=1, max_length=200)
    title: str | None = None
    mobile: str | None = None
    phone: str | None = None
    email: str | None = None
    wechat: str | None = None
    whatsapp: str | None = None
    responsible_products: str | None = None
    is_primary: bool = False
    notes: str | None = None

class SupplierProductIn(BaseModel):
    supplier_id: str
    product_name: str = Field(min_length=1, max_length=300)
    internal_keywords: list[str] = []
    nl_product_id: str | None = None
    nl_status: Literal["无 / 待确认", "已确认关联"] = "无 / 待确认"
    reference_model: str | None = None
    application: str | None = None
    technical_summary: str | None = None
    customizable: Literal["是", "否", "待确认"] = "待确认"
    sample_available: Literal["是", "否", "待确认"] = "待确认"
    capacity: str | None = None
    moq: str | None = None
    standard_lead_time: str | None = None
    packaging: str | None = None
    export_capacity: str = "待确认"
    notes: str | None = None

class SupplierFollowupIn(BaseModel):
    supplier_id: str
    supplier_project_link_id: str | None = None
    rfq_id: str | None = None
    date: str
    channel: Literal["邮件", "微信", "电话", "会议", "报价", "样品", "技术确认", "其他"] = "微信"
    content: str = Field(min_length=1, max_length=5000)
    conclusion: str | None = None
    next_action: str | None = None
    next_followup_date: str | None = None
    owner_name: str | None = None
    status: SupplierStatus = "待联系"
    create_task: bool = True

class SupplierProjectLinkIn(BaseModel):
    customer_id: str
    project_id: str
    supplier_id: str
    supplier_product_id: str | None = None
    customer_need: str | None = None
    reference_product: str | None = None
    match_status: Literal["待询价", "等资料", "技术评估", "已推荐", "已送样", "测试中", "已成交", "未匹配"] = "待询价"
    technical_match_notes: str | None = None
    quote_status: str | None = None
    sample_status: str | None = None
    current_risk: str | None = None
    next_action: str | None = None
    next_followup_date: str | None = None

class SupplierRfqIn(BaseModel):
    customer_id: str
    project_id: str
    supplier_id: str
    supplier_product_id: str | None = None
    demand_product: str = Field(min_length=1, max_length=300)
    reference_product: str | None = None
    end_application: str | None = None
    technical_requirements: str | None = None
    sample_quantity: str | None = None
    expected_monthly_usage: str | None = None
    expected_annual_usage: str | None = None
    destination_country: str | None = None
    requested_materials: list[str] | None = None
    status: Literal["草稿", "已发送", "供应商已回复", "技术评估", "关闭"] = "草稿"
    sent_date: str | None = None
    next_followup_date: str | None = None
    reply_content: str | None = None

class EmailStatusIn(BaseModel):
    status: Literal["unread", "new_lead", "linked", "followup_created", "completed"]

class EmailFollowupIn(BaseModel):
    content: str | None = Field(default=None, max_length=5000)
    next_action: str | None = Field(default=None, max_length=1000)

class EmailLinkIn(BaseModel):
    customer_id: str
    contact_name: str | None = Field(default=None, max_length=200)

class EmailCustomerCreateIn(CustomerIn):
    """Create a current-user customer from one reviewed mailbox message."""
    pass

class EmailCrmUpdateIn(BaseModel):
    customer_id: str
    project_id: str | None = None
    product_id: str | None = None
    customer_stage: str
    next_action: str = Field(min_length=1, max_length=1000)
    followup_date: str
    notes: str = Field(min_length=1, max_length=5000)
    create_task: bool = False
    task_date: str | None = None

class TaskIn(BaseModel):
    title: str = Field(min_length=1, max_length=200)
    description: str | None = Field(default=None, max_length=5000)
    category: Literal["外贸", "网站", "设计", "学习", "生活", "其他"] = "外贸"
    priority: Literal["important", "normal", "low"] = "normal"
    status: Literal["Pending", "Completed"] = "Pending"
    task_date: str
    start_time: str | None = None
    end_time: str | None = None
    estimated_minutes: int | None = Field(default=None, ge=5, le=1440)
    customer_id: str | None = None
    project_id: str | None = None
    product_id: str | None = None
    lead_id: str | None = None

class LeadSearchTaskIn(BaseModel):
    task_name: str = Field(min_length=1, max_length=200)
    product_keywords: list[str] = []
    application_keywords: list[str] = []
    target_countries: list[str] = []
    excluded_countries: list[str] = []
    target_company_types: list[str] = []
    source_urls: list[str] = []
    search_language: str = "English"
    max_results: int = Field(default=15, ge=1, le=100)
    daily_enabled: bool = False
    daily_run_time: str = "08:30"
    status: Literal["启用", "暂停"] = "启用"

class LeadReviewIn(BaseModel):
    status: Literal["待审核", "保留", "已排除", "已联系"]
    exclusion_reason: str | None = None
    notes: str | None = None
    watchlisted: bool | None = None

class LeadDevelopmentTaskIn(BaseModel):
    priority: Literal["important", "normal", "low"] = "normal"
    task_date: str
    suggested_next_action: str | None = None

class LeadConvertIn(BaseModel):
    email: str | None = None
    contact_person: str | None = None
    country: str | None = None
    product_interest: str | None = None
    application: str | None = None
    priority: Literal["HIGH", "MEDIUM HIGH", "MEDIUM"] = "MEDIUM"
    next_action: str | None = None
    next_followup_date: str | None = None
    notes: str | None = None
    customer_id: str | None = None

class TaskStatusIn(BaseModel):
    status: Literal["Pending", "Completed"]

class DailyLogIn(BaseModel):
    summary: str | None = Field(default=None, max_length=5000)
    problem: str | None = Field(default=None, max_length=5000)
    tomorrow_plan: str | None = Field(default=None, max_length=5000)
    rating: int | None = Field(default=None, ge=1, le=5)

class ImportPreviewIn(BaseModel):
    payload: dict[str, Any]

class ImportApplyIn(BaseModel):
    confirm_company_match: bool = False
    selected_customer_id: str | None = None

async def supabase(path: str, token: str, method: str = "GET", payload: dict[str, Any] | None = None) -> Any:
    cfg = settings()
    headers = {"apikey": cfg.supabase_anon_key, "Authorization": token, "Content-Type": "application/json", "Prefer": "return=representation"}
    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.request(method, f"{cfg.supabase_url}/rest/v1/{path}", headers=headers, json=payload)
    if response.status_code >= 400: raise HTTPException(response.status_code, response.text)
    return response.json() if response.content else None

def bearer(authorization: str | None) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing Supabase user token")
    return authorization

def is_internal_mail_address(address: str | None) -> bool:
    """Keep colleague-forwarded mail out of customer auto-matching."""
    value = (address or "").strip().lower()
    internal_addresses = {item.strip().lower() for item in settings().mail_internal_addresses.split(",") if item.strip()}
    internal_domains = {item.strip().lower() for item in settings().mail_internal_domains.split(",") if item.strip()}
    return value in internal_addresses or ("@" in value and value.rsplit("@", 1)[1] in internal_domains)

def _domain_from_url(value: str | None) -> str:
    return (urlparse(value or "").hostname or "").lower().removeprefix("www.")

def import_text(value: Any) -> str:
    return str(value or "").strip()

def import_key(value: Any) -> str:
    return import_text(value).casefold()

def import_nonempty(data: dict[str, Any]) -> dict[str, Any]:
    """Do not let blank fields from a ChatGPT summary erase CRM fields."""
    return {key: value for key, value in data.items() if value not in (None, "", [], {})}

def import_date(value: Any, field: str, *, required: bool = False) -> str | None:
    text = import_text(value)
    if not text:
        if required:
            raise HTTPException(422, f"{field} is required")
        return None
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise HTTPException(422, f"{field} must use YYYY-MM-DD") from exc

async def add_import_effect(
    token: str, batch_id: str, *, entity_type: str, action: str, record_id: str,
    before_data: dict[str, Any] | None, after_data: dict[str, Any] | None,
) -> None:
    await supabase("import_effects", token, "POST", {
        "import_batch_id": batch_id, "entity_type": entity_type, "action": action,
        "record_id": record_id, "before_data": before_data, "after_data": after_data,
    })

async def build_import_preview(token: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Validate a chat JSON packet and calculate its effects without writing CRM data."""
    if payload.get("schema_version") != "zhiwu-os-import/v1":
        raise HTTPException(422, "schema_version must be zhiwu-os-import/v1")
    if payload.get("intent") in {"upsert_supplier", "upsert_supplier_and_followup", "create_supplier_rfq", "link_supplier_to_customer_project"}:
        return await build_supplier_import_preview(token, payload)
    if payload.get("intent") != "upsert_customer_and_followup":
        raise HTTPException(422, "intent must be upsert_customer_and_followup")
    source = payload.get("source") if isinstance(payload.get("source"), dict) else {}
    import_date(source.get("date"), "source.date", required=True)
    customer = payload.get("customer") if isinstance(payload.get("customer"), dict) else {}
    match = payload.get("match") if isinstance(payload.get("match"), dict) else {}
    company_name = import_text(customer.get("company_name") or match.get("company_name"))
    email = import_text(customer.get("email") or match.get("customer_email")).lower()
    if not company_name and not email:
        raise HTTPException(422, "customer.company_name or customer.email is required")
    product_refs = payload.get("product_refs") if isinstance(payload.get("product_refs"), list) else []
    for index, product in enumerate(product_refs):
        if not isinstance(product, dict) or not import_text(product.get("code")).upper().startswith("NL-"):
            raise HTTPException(422, f"product_refs[{index}].code must start with NL-")

    customers = await supabase("customers?select=*&import_reverted=eq.false&archived_at=is.null&limit=500", token)
    products = await supabase("products?select=*&import_reverted=eq.false&archived_at=is.null&limit=500", token)
    email_matches = [row for row in customers if email and import_key(row.get("email")) == email]
    company_matches = [row for row in customers if company_name and import_key(row.get("company_name")) == import_key(company_name)]
    internal = is_internal_mail_address(email)
    candidates = [{"id": row["id"], "company_name": row.get("company_name"), "email": row.get("email")} for row in company_matches]
    if internal:
        customer_match = {"kind": "internal_forwarder", "customer_id": None, "candidates": candidates, "requires_confirmation": True, "message": "内部同事邮箱不能自动创建或自动匹配客户；请人工选择真实客户。"}
    elif email_matches:
        customer_match = {"kind": "email_exact", "customer_id": email_matches[0]["id"], "candidates": [], "requires_confirmation": False, "message": "按邮箱精确匹配到现有客户。"}
    elif len(company_matches) == 1:
        customer_match = {"kind": "company_manual_review", "customer_id": company_matches[0]["id"], "candidates": candidates, "requires_confirmation": True, "message": "按公司名称找到可能的现有客户，必须人工确认后才会更新。"}
    elif len(company_matches) > 1:
        customer_match = {"kind": "company_ambiguous", "customer_id": None, "candidates": candidates, "requires_confirmation": True, "message": "存在多个同名客户，请人工选择。"}
    else:
        customer_match = {"kind": "new_customer", "customer_id": None, "candidates": [], "requires_confirmation": True, "message": "未匹配到现有客户；确认后将新建客户。"}

    product_actions: list[dict[str, Any]] = []
    for product in product_refs:
        code = import_text(product.get("code")).upper()
        existing = next((row for row in products if import_key(row.get("product_code")) == import_key(code)), None)
        product_actions.append({"entity": "产品", "action": "更新" if existing else "新增", "label": f"{code} · {import_text(product.get('name')) or code}", "record_id": existing.get("id") if existing else None})

    actions = [{"entity": "客户", "action": "更新" if customer_match["kind"] in ("email_exact", "company_manual_review") else "新增", "label": company_name or email}]
    actions.extend(product_actions)
    project = payload.get("project") if isinstance(payload.get("project"), dict) else {}
    if import_text(project.get("name")):
        project_action = "待确认"
        if customer_match.get("customer_id"):
            project_rows = await supabase(f"projects?customer_id=eq.{customer_match['customer_id']}&import_reverted=eq.false&archived_at=is.null&select=id,project_name", token)
            existing_project = next((row for row in project_rows if import_key(row.get("project_name")) == import_key(project.get("name"))), None)
            project_action = "更新" if existing_project else "新增"
        actions.append({"entity": "项目", "action": project_action, "label": import_text(project.get("name"))})
    followup = payload.get("follow_up") if isinstance(payload.get("follow_up"), dict) else {}
    if import_text(followup.get("content")):
        actions.append({"entity": "跟进", "action": "新增", "label": import_text(followup.get("content"))[:100]})
    task = payload.get("task") if isinstance(payload.get("task"), dict) else {}
    if task.get("create") is True:
        actions.append({"entity": "每日任务", "action": "新增", "label": import_text(task.get("title")) or "待补充任务标题"})
    return {
        "customer_match": customer_match, "actions": actions,
        "requires_human_confirmation": bool(customer_match["requires_confirmation"] or (payload.get("review") or {}).get("requires_human_confirmation")),
        "uncertain_fields": (payload.get("review") or {}).get("uncertain_fields") or [],
    }

async def build_supplier_import_preview(token: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Review a supplier packet without treating a domestic contact as a CRM customer."""
    source = payload.get("source") if isinstance(payload.get("source"), dict) else {}
    import_date(source.get("date"), "source.date", required=True)
    supplier = payload.get("supplier") if isinstance(payload.get("supplier"), dict) else {}
    company_name = import_text(supplier.get("company_name") or (payload.get("match") or {}).get("company_name"))
    if not company_name:
        raise HTTPException(422, "supplier.company_name is required")
    contact = payload.get("contact") if isinstance(payload.get("contact"), dict) else {}
    email = import_key(supplier.get("main_email") or contact.get("email"))
    phone = import_key(supplier.get("main_phone") or contact.get("phone") or contact.get("mobile"))
    rows = await supabase("suppliers?select=*&archived_at=is.null&import_reverted=eq.false&limit=500", token)
    named = [row for row in rows if import_key(row.get("company_name")) == import_key(company_name)]
    exact = next((row for row in named if (email and import_key(row.get("main_email")) == email) or (phone and import_key(row.get("main_phone")) == phone)), None)
    candidates = [{"id": row["id"], "company_name": row.get("company_name"), "email": row.get("main_email")} for row in named]
    if exact:
        supplier_match = {"kind": "email_exact", "customer_id": exact["id"], "candidates": [], "requires_confirmation": False, "message": "按供应商公司名称和电话或邮箱匹配到现有档案。"}
    elif len(named) == 1:
        supplier_match = {"kind": "company_manual_review", "customer_id": named[0]["id"], "candidates": candidates, "requires_confirmation": True, "message": "按供应商公司名称找到可能档案；必须人工确认后才会更新。"}
    elif len(named) > 1:
        supplier_match = {"kind": "company_ambiguous", "customer_id": None, "candidates": candidates, "requires_confirmation": True, "message": "存在多个同名供应商，请先人工核对联系方式。"}
    else:
        supplier_match = {"kind": "new_supplier", "customer_id": None, "candidates": [], "requires_confirmation": True, "message": "未匹配到现有供应商；确认后将新增“待确认”供应商档案。"}
    intent = import_text(payload.get("intent"))
    actions = [{"entity": "供应商", "action": "更新" if supplier_match["kind"] in {"email_exact", "company_manual_review"} else "新增", "label": company_name}]
    if isinstance(payload.get("supplier_product"), dict):
        actions.append({"entity": "供应商产品", "action": "新增", "label": import_text(payload["supplier_product"].get("product_name")) or "待确认产品能力"})
    if intent == "upsert_supplier_and_followup" and isinstance(payload.get("follow_up"), dict):
        actions.append({"entity": "供应商跟进", "action": "新增", "label": import_text(payload["follow_up"].get("content"))[:100] or "供应商跟进"})
    if intent == "link_supplier_to_customer_project":
        actions.append({"entity": "供应链协同", "action": "新增", "label": import_text((payload.get("link") or {}).get("project_name")) or "关联客户项目"})
    if intent == "create_supplier_rfq":
        actions.append({"entity": "供应商 RFQ", "action": "新增", "label": import_text((payload.get("rfq") or {}).get("demand_product")) or "待确认询价需求"})
    uncertain = list((payload.get("review") or {}).get("uncertain_fields") or [])
    if import_text(supplier.get("supplier_type")) not in {"工厂", "贸易商"}: uncertain.append("supplier_type 未经确认，写入时将保持“待确认”。")
    if import_text(supplier.get("export_status")) not in {"可出口", "不可出口"}: uncertain.append("出口能力没有正式资料，写入时将保持“待确认”。")
    return {"supplier_match": supplier_match, "actions": actions, "requires_human_confirmation": True, "uncertain_fields": uncertain}

async def record_timeline_event(
    token: str, *, title: str, event_type: Literal["task", "email", "crm", "project", "note"],
    source: str, related_id: str | None = None, customer_id: str | None = None,
    project_id: str | None = None, product_id: str | None = None,
    supplier_id: str | None = None, supplier_rfq_id: str | None = None,
    event_date: str | None = None, event_time: str | None = None,
) -> None:
    """A failed optional timeline write must not block an existing CRM/mail action."""
    try:
        await supabase("timeline_events", token, "POST", {
            "title": title, "event_type": event_type, "source": source, "related_id": related_id,
            "customer_id": customer_id, "project_id": project_id, "product_id": product_id,
            "supplier_id": supplier_id, "supplier_rfq_id": supplier_rfq_id,
            "event_date": event_date or str(date.today()),
            "event_time": event_time or datetime.now().strftime("%H:%M:%S"),
        })
    except HTTPException:
        pass

DEMO_PRODUCTS = [
    {"product_name": "NL-007", "product_code": "NL-007", "category": "Barrier Masterbatch", "application": "PPC Film", "description": "Food packaging barrier masterbatch", "notes": "Active"},
    {"product_name": "NL-PHA-21", "product_code": "NL-PHA-21", "category": "Water-based Barrier Coating", "application": "Paper Packaging", "description": "PFAS-free water-based barrier coating", "notes": "Active"},
    {"product_name": "HM-800", "product_code": "HM-800", "category": "Bio-based Polyester Plasticizer", "application": "PVC Film", "description": "Bio-based polyester plasticizer", "notes": "Active"},
    {"product_name": "ESO", "product_code": "ESO", "category": "Epoxidized Soybean Oil", "application": "PVC Plasticizer", "description": "Epoxidized soybean oil", "notes": "Active"},
    {"product_name": "MCPP", "product_code": "MCPP", "category": "Maleic Anhydride Modified Chlorinated Polypropylene", "application": "Adhesion Promoter", "description": "Adhesion promoter for PP", "notes": "Active"},
]
DEMO_CUSTOMERS = [
    {"company_name": "Uflex", "country": "India", "contact_person": "Dileep", "email": "dileep@uflex.co.in", "whatsapp": "+91 98 221 8608", "product_interest": "NL-007", "application": "PPC Film / Food Packaging", "customer_stage": "Sample Payment Pending", "priority": "HIGH", "status_label": "Waiting sample payment", "status_tone": "warning", "last_contact_date": "2026-08-20", "next_followup_date": "2026-08-25", "notes": "25kg sample · USD 310"},
    {"company_name": "Agrileaf", "country": "India", "contact_person": "Vaibhav", "email": "vaibhav@agrileaf.in", "whatsapp": "+91 97 552 3901", "product_interest": "NL-PHA-21", "application": "Water-based Barrier Coating", "customer_stage": "Sample Payment", "priority": "HIGH", "status_label": "Waiting payment", "status_tone": "warning", "last_contact_date": "2026-08-18", "next_followup_date": "2026-08-24", "notes": "5kg sample · USD 150"},
    {"company_name": "Flexo", "country": "Philippines", "contact_person": "Joselito", "email": "joselito@flexo.ph", "whatsapp": "+63 917 555 0190", "product_interest": "E4050 Replacement Project", "application": "Glassine Extrusion Coating", "customer_stage": "Technical Testing", "priority": "HIGH", "status_label": "Waiting customer sample", "status_tone": "attention", "last_contact_date": "2026-08-20", "next_followup_date": "2026-08-26", "notes": "Henkel Proxmelt E4050 replacement"},
    {"company_name": "FLEX Design", "country": "Netherlands", "contact_person": "Dominic", "email": "dominic@flexdesign.nl", "whatsapp": "+31 6 1890 3033", "product_interest": "NL-PHA-21", "application": "PFAS-free Paper Cup Barrier Coating", "customer_stage": "Technical Confirmation", "priority": "MEDIUM HIGH", "status_label": "Waiting sample confirmation", "status_tone": "warning", "last_contact_date": "2026-08-19", "next_followup_date": "2026-08-27", "notes": "Spray / dip coating"},
    {"company_name": "ATSajan", "country": "Thailand", "contact_person": "Anchasa", "email": "anchasa@atsajan.co.th", "whatsapp": "+66 81 553 2871", "product_interest": "MCPP", "application": "Polypropylene Adhesion Modification", "customer_stage": "Maintain Relationship", "priority": "MEDIUM", "status_label": "Maintain relationship", "status_tone": "success", "last_contact_date": "2026-08-16", "next_followup_date": "2026-09-02", "notes": "Potential demand: 20 tons/year"},
    {"company_name": "Inkofix", "country": "India", "contact_person": "LN Garg", "email": "lngarg@inkofix.in", "whatsapp": "+91 99 871 1640", "product_interest": "NL-PHA-21", "application": "Water-based Barrier Coating", "customer_stage": "Quotation", "priority": "MEDIUM HIGH", "status_label": "Price discussion", "status_tone": "attention", "last_contact_date": "2026-08-15", "next_followup_date": "2026-08-28", "notes": "USD 5,550/T CIF Mundra"},
]

@app.post("/api/auth/login")
async def login(credentials: LoginIn):
    """The browser never receives a Supabase service-role key."""
    cfg = settings()
    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.post(
            f"{cfg.supabase_url}/auth/v1/token?grant_type=password",
            headers={"apikey": cfg.supabase_anon_key, "Content-Type": "application/json"},
            json=credentials.model_dump(),
        )
    if response.status_code >= 400:
        raise HTTPException(401, "Invalid login")
    return response.json()

@app.post("/api/auth/update-password")
async def update_password(payload: PasswordIn, authorization: str | None = Header(default=None)):
    """Change the password for the currently authenticated Supabase user."""
    cfg = settings()
    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.put(
            f"{cfg.supabase_url}/auth/v1/user",
            headers={"apikey": cfg.supabase_anon_key, "Authorization": bearer(authorization), "Content-Type": "application/json"},
            json=payload.model_dump(),
        )
    if response.status_code >= 400:
        raise HTTPException(response.status_code, "Password update failed. Please log in again and retry.")
    return {"updated": True}

@app.post("/api/auth/recover")
async def request_password_recovery(payload: RecoveryIn):
    """Send a Supabase password recovery link without exposing account existence."""
    cfg = settings()
    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.post(
            f"{cfg.supabase_url}/auth/v1/recover",
            headers={"apikey": cfg.supabase_anon_key, "Content-Type": "application/json"},
            json={"email": payload.email, "redirect_to": cfg.public_app_url},
        )
    if response.status_code >= 400:
        raise HTTPException(response.status_code, "Unable to send password recovery email")
    return {"sent": True}

@app.get("/health")
def health(): return {"status": "ok"}

@app.post("/api/demo/seed")
async def seed_demo(authorization: str | None = Header(default=None)):
    """Initialize one authenticated workspace with the V1.1 trade CRM demo dataset."""
    token = bearer(authorization)
    existing = await supabase("customers?archived_at=is.null&select=id&limit=1", token)
    if existing:
        return {"seeded": False, "reason": "workspace already has customers"}
    product_rows = await supabase("products", token, "POST", DEMO_PRODUCTS)
    customer_rows = await supabase("customers", token, "POST", DEMO_CUSTOMERS)
    product_ids = {row["product_code"]: row["id"] for row in product_rows}
    customer_ids = {row["company_name"]: row["id"] for row in customer_rows}
    await supabase("projects", token, "POST", [
        {"customer_id": customer_ids["Uflex"], "project_name": "PPC Film Barrier Masterbatch", "product_id": product_ids["NL-007"], "application": "Food Packaging", "stage": "Sample Payment Pending", "notes": "25kg sample"},
        {"customer_id": customer_ids["Agrileaf"], "project_name": "Water-based Barrier Coating Sample", "product_id": product_ids["NL-PHA-21"], "application": "Water-based Barrier Coating", "stage": "Sample Payment", "notes": "5kg sample"},
        {"customer_id": customer_ids["Flexo"], "project_name": "E4050 Replacement Project", "product_id": None, "application": "Glassine Extrusion Coating", "stage": "Technical Testing", "notes": "Prepare 2kg test sample"},
        {"customer_id": customer_ids["FLEX Design"], "project_name": "PFAS-free Paper Cup Barrier Coating", "product_id": product_ids["NL-PHA-21"], "application": "Paper Cup Barrier Coating", "stage": "Technical Confirmation", "notes": "Prepare 1–2L sample"},
        {"customer_id": customer_ids["ATSajan"], "project_name": "PP Adhesion Modification", "product_id": product_ids["MCPP"], "application": "Polypropylene Adhesion Modification", "stage": "Maintain Relationship", "notes": "20 tons/year potential"},
        {"customer_id": customer_ids["Inkofix"], "project_name": "Mundra Quotation Project", "product_id": product_ids["NL-PHA-21"], "application": "Water-based Barrier Coating", "stage": "Quotation", "notes": "CIF Mundra price discussion"},
    ])
    await supabase("followups", token, "POST", [
        {"customer_id": customer_ids["Uflex"], "date": "2026-08-20", "content": "Confirmed performance targets and 25kg sample requirement", "next_action": "Send PI and confirm payment"},
        {"customer_id": customer_ids["Agrileaf"], "date": "2026-08-18", "content": "Customer provided delivery address", "next_action": "Confirm USD 150 payment"},
        {"customer_id": customer_ids["Flexo"], "date": "2026-08-20", "content": "Sent alternative product information", "next_action": "Receive customer sample"},
    ])
    await supabase("quotes", token, "POST", [
        {"customer_id": customer_ids["Uflex"], "product_id": product_ids["NL-007"], "quantity": "25kg", "amount": 310, "currency": "USD", "trade_term": "Sample + Express", "status": "Pending Payment"},
        {"customer_id": customer_ids["Inkofix"], "product_id": product_ids["NL-PHA-21"], "quantity": "18 tons / 20GP", "amount": 5550, "currency": "USD", "trade_term": "CIF Mundra / T", "status": "Price Discussion"},
    ])
    return {"seeded": True, "customers": len(customer_rows), "products": len(product_rows)}

@app.get("/api/customers")
async def list_customers(authorization: str | None = Header(default=None), limit: int = Query(100, le=100)):
    return await supabase(f"customers?select=*&import_reverted=eq.false&archived_at=is.null&order=created_at.desc&limit={limit}", bearer(authorization))

@app.get("/api/workspace-members")
async def list_workspace_members(authorization: str | None = Header(default=None)):
    """Read-only labels for customer ownership; RLS remains the source of truth."""
    return await supabase("workspace_members?select=user_id,display_name,role&order=display_name.asc", bearer(authorization))

@app.post("/api/customers", status_code=201)
async def create_customer(customer: CustomerIn, authorization: str | None = Header(default=None)):
    return await supabase("customers", bearer(authorization), "POST", customer.model_dump(exclude_none=True))

@app.patch("/api/customers/{customer_id}")
async def update_customer(customer_id: str, customer: CustomerIn, authorization: str | None = Header(default=None)):
    return await supabase(f"customers?id=eq.{customer_id}&archived_at=is.null", bearer(authorization), "PATCH", customer.model_dump(exclude_none=True))

@app.get("/api/products")
async def list_products(authorization: str | None = Header(default=None)):
    return await supabase("products?select=*&import_reverted=eq.false&archived_at=is.null&order=product_name.asc", bearer(authorization))

@app.post("/api/products", status_code=201)
async def create_product(product: ProductIn, authorization: str | None = Header(default=None)):
    return await supabase("products", bearer(authorization), "POST", product.model_dump(exclude_none=True))

@app.get("/api/product-customer-relations")
async def list_product_customer_relations(authorization: str | None = Header(default=None)):
    return await supabase("product_customer_relations?select=*&import_reverted=eq.false&archived_at=is.null&order=created_at.desc", bearer(authorization))

@app.post("/api/products/{product_id}/customers", status_code=201)
async def link_product_to_customer(product_id: str, payload: ProductCustomerRelationIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    product_rows = await supabase(f"products?id=eq.{product_id}&archived_at=is.null&select=id", token)
    customer_rows = await supabase(f"customers?id=eq.{payload.customer_id}&archived_at=is.null&select=id", token)
    if not product_rows or not customer_rows:
        raise HTTPException(404, "Product or customer not found")
    existing = await supabase(f"product_customer_relations?product_id=eq.{product_id}&customer_id=eq.{payload.customer_id}&import_reverted=eq.false&archived_at=is.null&select=*", token)
    if existing:
        return existing[0]
    rows = await supabase("product_customer_relations", token, "POST", {"product_id": product_id, "customer_id": payload.customer_id})
    return rows[0]

@app.get("/api/followups")
async def list_followups(authorization: str | None = Header(default=None)):
    return await supabase("followups?select=*&import_reverted=eq.false&archived_at=is.null&order=date.desc", bearer(authorization))

@app.post("/api/followups", status_code=201)
async def create_followup(followup: FollowupIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    rows = await supabase("followups", token, "POST", followup.model_dump(exclude_none=True))
    await record_timeline_event(token, title=f"跟进客户：{followup.content}", event_type="crm", source="followup", related_id=rows[0]["id"], customer_id=followup.customer_id, event_date=followup.date)
    return rows

@app.get("/api/projects")
async def list_projects(authorization: str | None = Header(default=None)):
    return await supabase("projects?select=*&import_reverted=eq.false&archived_at=is.null&order=created_at.desc", bearer(authorization))

@app.post("/api/projects", status_code=201)
async def create_project(project: ProjectIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    rows = await supabase("projects", token, "POST", project.model_dump(exclude_none=True))
    await record_timeline_event(token, title=f"创建项目：{project.project_name}", event_type="project", source="project", related_id=rows[0]["id"], customer_id=project.customer_id, product_id=project.product_id)
    return rows

@app.patch("/api/projects/{project_id}")
async def update_project(project_id: str, project: ProjectUpdateIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    existing = await supabase(f"projects?id=eq.{project_id}&archived_at=is.null&select=id,customer_id", token)
    if not existing:
        raise HTTPException(404, "Project not found")
    rows = await supabase(f"projects?id=eq.{project_id}", token, "PATCH", project.model_dump(exclude_none=True))
    record = rows[0]
    await record_timeline_event(token, title=f"更新项目：{record['project_name']}", event_type="project", source="project", related_id=record["id"], customer_id=existing[0]["customer_id"], product_id=record.get("product_id"))
    return rows

@app.get("/api/tasks")
async def list_tasks(
    authorization: str | None = Header(default=None), task_date: str | None = None,
    from_date: str | None = Query(default=None), to_date: str | None = Query(default=None),
):
    filters = ["select=*", "import_reverted=eq.false", "archived_at=is.null", "order=task_date.asc,start_time.asc"]
    if task_date:
        filters.append(f"task_date=eq.{task_date}")
    if from_date:
        filters.append(f"task_date=gte.{from_date}")
    if to_date:
        filters.append(f"task_date=lte.{to_date}")
    return await supabase(f"tasks?{'&'.join(filters)}", bearer(authorization))

@app.post("/api/tasks", status_code=201)
async def create_task(task: TaskIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    rows = await supabase("tasks", token, "POST", task.model_dump(exclude_none=True))
    record = rows[0]
    await record_timeline_event(token, title=f"计划：{record['title']}", event_type="task", source="task", related_id=record["id"], customer_id=record.get("customer_id"), project_id=record.get("project_id"), product_id=record.get("product_id"), event_date=record["task_date"], event_time=record.get("start_time"))
    return record

@app.patch("/api/tasks/{task_id}")
async def update_task_status(task_id: str, payload: TaskStatusIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    existing = await supabase(f"tasks?id=eq.{task_id}&archived_at=is.null&select=*", token)
    if not existing:
        raise HTTPException(404, "Task not found")
    task = existing[0]
    update = {"status": payload.status, "completed_at": datetime.now().isoformat() if payload.status == "Completed" else None}
    rows = await supabase(f"tasks?id=eq.{task_id}", token, "PATCH", update)
    if payload.status == "Completed":
        await record_timeline_event(token, title=f"完成任务：{task['title']}", event_type="task", source="task", related_id=task_id, customer_id=task.get("customer_id"), project_id=task.get("project_id"), product_id=task.get("product_id"))
    return rows[0]

@app.get("/api/lead-search-tasks")
async def list_lead_search_tasks(authorization: str | None = Header(default=None)):
    return await supabase("lead_search_tasks?deleted_at=is.null&select=*&order=created_at.asc", bearer(authorization))

@app.post("/api/lead-search-tasks", status_code=201)
async def create_lead_search_task(payload: LeadSearchTaskIn, authorization: str | None = Header(default=None)):
    rows = await supabase("lead_search_tasks", bearer(authorization), "POST", payload.model_dump())
    return rows[0]

@app.patch("/api/lead-search-tasks/{task_id}")
async def update_lead_search_task(task_id: str, payload: LeadSearchTaskIn, authorization: str | None = Header(default=None)):
    # Keep existing V1.13 tasks usable while V1.14's optional directory-source
    # column is being rolled out. Explicit non-default sources still persist.
    rows = await supabase(f"lead_search_tasks?id=eq.{task_id}&deleted_at=is.null", bearer(authorization), "PATCH", {**payload.model_dump(exclude_defaults=True), "updated_at": datetime.now().isoformat()})
    if not rows: raise HTTPException(404, "Lead search task not found")
    return rows[0]

@app.delete("/api/lead-search-tasks/{task_id}")
async def delete_lead_search_task(task_id: str, authorization: str | None = Header(default=None)):
    """Soft-delete configuration only; discovered leads and run history remain."""
    token = bearer(authorization)
    active = await supabase(f"lead_discovery_runs?task_id=eq.{task_id}&status=eq.%E8%BF%90%E8%A1%8C%E4%B8%AD&select=id&limit=1", token)
    if active:
        raise HTTPException(409, "该任务正在运行，请等待完成后再删除")
    rows = await supabase(f"lead_search_tasks?id=eq.{task_id}&deleted_at=is.null", token, "PATCH", {"status": "暂停", "daily_enabled": False, "deleted_at": datetime.now().isoformat(), "updated_at": datetime.now().isoformat()})
    if not rows: raise HTTPException(404, "Lead search task not found")
    return {"deleted": True, "task_id": task_id, "message": "任务已删除；已有线索与运行记录已保留。"}

@app.get("/api/lead-discovery-runs")
async def list_lead_discovery_runs(authorization: str | None = Header(default=None), limit: int = Query(50, le=100)):
    return await supabase(f"lead_discovery_runs?select=*&order=started_at.desc&limit={limit}", bearer(authorization))

async def _run_lead_task_for_user(token: str, task_id: str, trigger: str) -> dict[str, Any]:
    task_rows = await supabase(f"lead_search_tasks?id=eq.{task_id}&deleted_at=is.null&select=*&limit=1", token)
    if not task_rows: raise HTTPException(404, "Lead search task not found")
    from .lead_discovery import RestStore, run_task_once
    try:
        return await run_task_once(RestStore(token), task_rows[0], trigger)
    except Exception as exc:
        raise HTTPException(502, f"公开网页搜索失败：{str(exc)[:500]}")

@app.post("/api/lead-search-tasks/{task_id}/run")
async def run_lead_search_task(task_id: str, background_tasks: BackgroundTasks, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    active = await supabase(f"lead_discovery_runs?task_id=eq.{task_id}&status=eq.%E8%BF%90%E8%A1%8C%E4%B8%AD&select=id&limit=1", token)
    if active: return {"run_id": active[0]["id"], "status": "运行中", "message": "该任务正在按合规限速运行。"}
    background_tasks.add_task(_run_lead_task_for_user, token, task_id, "manual")
    return {"status": "已开始", "message": "已在服务器后台开始公开网页搜索；完成后刷新即可查看审核池和运行日志。"}

async def _run_enabled_lead_tasks(token: str) -> None:
    tasks = await supabase("lead_search_tasks?deleted_at=is.null&status=eq.%E5%90%AF%E7%94%A8&select=id&order=created_at.asc", token)
    for task in tasks:
        try: await _run_lead_task_for_user(token, task["id"], "manual")
        except HTTPException: continue

@app.post("/api/lead-search-tasks/run-enabled")
async def run_enabled_lead_search_tasks(background_tasks: BackgroundTasks, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    background_tasks.add_task(_run_enabled_lead_tasks, token)
    return {"status": "已开始", "message": "所有启用任务将在服务器后台按顺序运行。"}

@app.get("/api/customer-leads")
async def list_customer_leads(authorization: str | None = Header(default=None), limit: int = Query(300, le=500)):
    return await supabase(f"customer_leads?select=*&order=discovered_at.desc&limit={limit}", bearer(authorization))

@app.patch("/api/customer-leads/{lead_id}")
async def review_customer_lead(lead_id: str, payload: LeadReviewIn, authorization: str | None = Header(default=None)):
    if payload.status != "已排除" and payload.exclusion_reason:
        raise HTTPException(422, "仅“已排除”线索可保存排除原因")
    data = {**payload.model_dump(exclude_none=True), "updated_at": datetime.now().isoformat()}
    if payload.status == "已排除":
        data.update({"verification_bucket": "排除名单", "verification_conclusion": payload.exclusion_reason or "人工排除"})
    rows = await supabase(f"customer_leads?id=eq.{lead_id}", bearer(authorization), "PATCH", data)
    if not rows: raise HTTPException(404, "Lead not found")
    return rows[0]

@app.post("/api/customer-leads/{lead_id}/development-task", status_code=201)
async def create_development_task(lead_id: str, payload: LeadDevelopmentTaskIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    lead_rows = await supabase(f"customer_leads?id=eq.{lead_id}&select=*&limit=1", token)
    if not lead_rows: raise HTTPException(404, "Lead not found")
    lead = lead_rows[0]
    if lead.get("verification_bucket") != "严格客户名单":
        raise HTTPException(422, "仅通过严格客户核验的企业可以创建开发任务")
    task_rows = await supabase("tasks", token, "POST", {
        "title": f"研究 / 联系 {lead['company_name']}", "description": payload.suggested_next_action or f"查看官网并确认是否有 {', '.join((lead.get('discovered_application_keywords') or ['目标应用'])[:2])} 业务；仅在人工确认后决定是否联系。",
        "category": "外贸", "priority": payload.priority, "status": "Pending", "task_date": payload.task_date, "lead_id": lead_id,
    })
    task = task_rows[0]
    await supabase(f"customer_leads?id=eq.{lead_id}", token, "PATCH", {"development_task_id": task["id"], "status": "保留", "updated_at": datetime.now().isoformat()})
    return task

@app.post("/api/customer-leads/{lead_id}/convert")
async def convert_customer_lead(lead_id: str, payload: LeadConvertIn, authorization: str | None = Header(default=None)):
    """Explicit user review is required before any CRM write; no NL code is invented."""
    token = bearer(authorization)
    lead_rows = await supabase(f"customer_leads?id=eq.{lead_id}&select=*&limit=1", token)
    if not lead_rows: raise HTTPException(404, "Lead not found")
    lead = lead_rows[0]
    if lead.get("verification_bucket") != "严格客户名单":
        raise HTTPException(422, "仅通过严格客户核验的企业可以载入 CRM")
    email = (payload.email or lead.get("public_business_email") or "").strip().lower()
    identity = f"{lead.get('company_name','')} {payload.contact_person or lead.get('public_contact_name') or ''}".lower()
    if is_internal_mail_address(email) or any(name in identity for name in ("zhiwu", "peter")):
        raise HTTPException(422, "内部同事不能转为海外客户")
    customers = await supabase("customers?select=*&archived_at=is.null&import_reverted=eq.false&limit=500", token)
    lead_domain = _domain_from_url(lead.get("website"))
    existing = next((row for row in customers if payload.customer_id == row.get("id")), None)
    if not existing:
        existing = next((row for row in customers if email and (row.get("email") or "").lower() == email), None)
    if not existing:
        existing = next((row for row in customers if lead_domain and _domain_from_url(row.get("website")) == lead_domain), None)
    if not existing:
        existing = next((row for row in customers if (row.get("company_name") or "").strip().lower() == (lead.get("company_name") or "").strip().lower()), None)
    patch = {
        "contact_person": payload.contact_person or lead.get("public_contact_name"), "country": payload.country or lead.get("country"),
        "website": lead.get("website"), "product_interest": payload.product_interest,
        "application": payload.application or lead.get("possible_need"), "priority": payload.priority,
        "next_action": [payload.next_action] if payload.next_action else None,
        "next_followup_date": payload.next_followup_date, "notes": payload.notes,
    }
    patch = {key: value for key, value in patch.items() if value not in (None, "")}
    if existing:
        rows = await supabase(f"customers?id=eq.{existing['id']}", token, "PATCH", patch)
        customer = rows[0]
        action = "updated"
    else:
        if not email:
            raise HTTPException(422, "新建 CRM 客户需要公开商务邮箱；请先补充邮箱或仅保留为线索")
        customer_data = {"company_name": lead["company_name"], "country": payload.country or lead.get("country") or "待确认", "contact_person": payload.contact_person or lead.get("public_contact_name") or "待确认", "email": email, "customer_stage": "New", **patch}
        rows = await supabase("customers", token, "POST", customer_data)
        customer = rows[0]
        action = "created"
    await record_timeline_event(token, title=f"线索来源：{lead['company_name']}（公开网页审核转入）", event_type="crm", source="lead_discovery", related_id=lead_id, customer_id=customer["id"])
    await supabase(f"customer_leads?id=eq.{lead_id}", token, "PATCH", {"status": "已转 CRM", "crm_customer_id": customer["id"], "converted_at": datetime.now().isoformat(), "updated_at": datetime.now().isoformat()})
    return {"customer": customer, "action": action}

@app.get("/api/customer-leads/strict-export")
async def export_strict_customer_leads(authorization: str | None = Header(default=None)):
    """Export only fully verified, CRM-eligible companies as a real .xlsx file."""
    rows = await supabase("customer_leads?select=*&verification_bucket=eq.%E4%B8%A5%E6%A0%BC%E5%AE%A2%E6%88%B7%E5%90%8D%E5%8D%95&order=discovered_at.desc&limit=500", bearer(authorization))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "严格客户名单"
    headings = ["公司名称", "国家/地区", "企业类型", "官网", "联系人/部门", "公开邮箱", "公开电话", "公开地址/经营范围", "产品/应用证据摘要", "应用范围", "潜在匹配点", "推荐首联部门", "首轮需确认问题", "真实性核验结论", "匹配等级", "官网来源", "联系方式来源", "产品证据来源", "抓取时间"]
    sheet.append(headings)
    for cell in sheet[1]: cell.font = Font(bold=True)
    for row in rows:
        values = [row.get("company_name"), row.get("country"), row.get("company_type"), row.get("official_homepage_url") or row.get("website"), row.get("public_contact_name") or row.get("contact_department"), row.get("public_business_email"), row.get("public_business_phone"), row.get("official_address") or row.get("business_scope"), row.get("product_evidence_summary"), ", ".join(row.get("discovered_application_keywords") or []), row.get("possible_need"), row.get("recommended_contact_department"), row.get("first_contact_questions"), row.get("verification_conclusion"), row.get("matching_grade"), row.get("company_source_url"), row.get("contact_source_url"), row.get("product_evidence_url"), row.get("verified_at") or row.get("discovered_at")]
        sheet.append(values)
        excel_row = sheet.max_row
        for index in (4, 16, 17, 18):
            value = sheet.cell(excel_row, index).value
            if value:
                sheet.cell(excel_row, index).hyperlink = str(value)
                sheet.cell(excel_row, index).style = "Hyperlink"
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 14), 48)
    stream = BytesIO(); workbook.save(stream)
    return Response(stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=strict-customers.xlsx"})

@app.get("/api/daily-logs")
async def get_daily_log(log_date: str, authorization: str | None = Header(default=None)):
    rows = await supabase(f"daily_logs?log_date=eq.{log_date}&select=*&limit=1", bearer(authorization))
    return rows[0] if rows else None

@app.put("/api/daily-logs/{log_date}")
async def save_daily_log(log_date: str, payload: DailyLogIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    existing = await supabase(f"daily_logs?log_date=eq.{log_date}&select=id&limit=1", token)
    data = {**payload.model_dump(exclude_none=True), "updated_at": datetime.now().isoformat()}
    if existing:
        rows = await supabase(f"daily_logs?id=eq.{existing[0]['id']}", token, "PATCH", data)
    else:
        rows = await supabase("daily_logs", token, "POST", {**data, "log_date": log_date})
    await record_timeline_event(token, title="完成今日复盘", event_type="note", source="daily_log", related_id=rows[0]["id"], event_date=log_date)
    return rows[0]

@app.get("/api/timeline")
async def list_timeline(
    authorization: str | None = Header(default=None), event_date: str | None = None,
    from_date: str | None = Query(default=None), to_date: str | None = Query(default=None), limit: int = Query(300, le=500),
):
    filters = ["select=*", "import_reverted=eq.false", "archived_at=is.null", "order=event_date.desc,event_time.desc", f"limit={limit}"]
    if event_date:
        filters.append(f"event_date=eq.{event_date}")
    if from_date:
        filters.append(f"event_date=gte.{from_date}")
    if to_date:
        filters.append(f"event_date=lte.{to_date}")
    return await supabase(f"timeline_events?{'&'.join(filters)}", bearer(authorization))

@app.get("/api/imports")
async def list_import_batches(authorization: str | None = Header(default=None), limit: int = Query(20, le=100)):
    return await supabase(f"import_batches?select=id,schema_version,source_type,source_date,source_reference,preview,status,applied_at,reverted_at,created_at&order=created_at.desc&limit={limit}", bearer(authorization))

@app.post("/api/imports/preview", status_code=201)
async def preview_import(payload: ImportPreviewIn, authorization: str | None = Header(default=None)):
    """Create a draft import batch. This endpoint never touches CRM business records."""
    token = bearer(authorization)
    preview = await build_import_preview(token, payload.payload)
    source = payload.payload.get("source") or {}
    rows = await supabase("import_batches", token, "POST", {
        "schema_version": payload.payload["schema_version"],
        "source_type": import_text(source.get("type")) or "chat_summary",
        "source_date": import_date(source.get("date"), "source.date", required=True),
        "source_reference": import_text(source.get("reference")),
        "raw_payload": payload.payload, "preview": preview, "status": "draft",
    })
    return {"batch": rows[0], "preview": preview}

async def apply_supplier_import(token: str, batch_id: str, payload: dict[str, Any], confirmation: ImportApplyIn) -> dict[str, Any]:
    preview = await build_supplier_import_preview(token, payload)
    match = preview["supplier_match"]
    if match["kind"] == "company_ambiguous":
        raise HTTPException(422, "同名供应商存在歧义；请先在供应商中心人工核对后再导入")
    if match["kind"] == "company_manual_review" and not confirmation.confirm_company_match:
        raise HTTPException(422, "供应商公司名称匹配必须人工确认后才能更新")
    source = payload.get("source") or {}; source_date = import_date(source.get("date"), "source.date", required=True) or str(date.today())
    supplier_data = payload.get("supplier") or {}; contact_data = payload.get("contact") if isinstance(payload.get("contact"), dict) else {}
    supplier_type = import_text(supplier_data.get("supplier_type")); supplier_type = supplier_type if supplier_type in {"工厂", "贸易商"} else "待确认"
    export_status = import_text(supplier_data.get("export_status")); export_status = export_status if export_status in {"可出口", "不可出口"} else "待确认"
    values = {
        "company_name": import_text(supplier_data.get("company_name") or (payload.get("match") or {}).get("company_name")),
        "english_name": import_text(supplier_data.get("english_name")) or None, "country": import_text(supplier_data.get("country")) or "China",
        "province": import_text(supplier_data.get("province")) or None, "city": import_text(supplier_data.get("city")) or None,
        "address": import_text(supplier_data.get("address")) or None, "website": import_text(supplier_data.get("website")) or None,
        "supplier_type": supplier_type, "export_status": export_status,
        "main_phone": import_text(supplier_data.get("main_phone") or contact_data.get("phone") or contact_data.get("mobile")) or None,
        "main_email": import_text(supplier_data.get("main_email") or contact_data.get("email")).lower() or None,
        "wechat": import_text(supplier_data.get("wechat") or contact_data.get("wechat")) or None,
        "product_keywords": supplier_data.get("product_keywords") if isinstance(supplier_data.get("product_keywords"), list) else [],
        "product_categories": supplier_data.get("product_categories") if isinstance(supplier_data.get("product_categories"), list) else [],
        "supplier_tags": supplier_data.get("tags") if isinstance(supplier_data.get("tags"), list) else [],
        "current_status": import_text(supplier_data.get("current_status")) or "待联系", "last_contact_date": source_date,
        "next_action": import_text(supplier_data.get("next_action")) or None, "next_followup_date": import_date(supplier_data.get("next_followup_date"), "supplier.next_followup_date"),
        "notes": import_text(supplier_data.get("notes")) or None, "risk_notes": import_text(supplier_data.get("risk_notes")) or None,
        "updated_at": datetime.now().isoformat(),
    }
    if not values["company_name"]: raise HTTPException(422, "supplier.company_name is required")
    if match.get("customer_id"):
        current = (await supabase(f"suppliers?id=eq.{match['customer_id']}&select=*&limit=1", token))[0]
        changed = import_nonempty(values); before = {key: current.get(key) for key in changed}
        supplier = (await supabase(f"suppliers?id=eq.{current['id']}", token, "PATCH", changed))[0]
        await add_import_effect(token, batch_id, entity_type="supplier", action="updated", record_id=supplier["id"], before_data=before, after_data=changed); supplier_action = "updated"
    else:
        supplier = (await supabase("suppliers", token, "POST", {**values, "import_batch_id": batch_id}))[0]
        await add_import_effect(token, batch_id, entity_type="supplier", action="created", record_id=supplier["id"], before_data=None, after_data=values); supplier_action = "created"
    contact = None
    if import_text(contact_data.get("name")):
        contact_values = {"supplier_id": supplier["id"], "name": import_text(contact_data.get("name")), "title": import_text(contact_data.get("title")) or None, "mobile": import_text(contact_data.get("mobile")) or None, "phone": import_text(contact_data.get("phone")) or None, "email": import_text(contact_data.get("email")) or None, "wechat": import_text(contact_data.get("wechat")) or None, "whatsapp": import_text(contact_data.get("whatsapp")) or None, "responsible_products": import_text(contact_data.get("responsible_products")) or None, "is_primary": bool(contact_data.get("is_primary")), "notes": import_text(contact_data.get("notes")) or None, "import_batch_id": batch_id}
        contact = (await supabase("supplier_contacts", token, "POST", contact_values))[0]
        await add_import_effect(token, batch_id, entity_type="supplier_contact", action="created", record_id=contact["id"], before_data=None, after_data=contact_values)
    supplier_product = None; product_data = payload.get("supplier_product") if isinstance(payload.get("supplier_product"), dict) else {}
    if import_text(product_data.get("product_name")):
        product_values = {"supplier_id": supplier["id"], "product_name": import_text(product_data.get("product_name")), "internal_keywords": product_data.get("internal_keywords") if isinstance(product_data.get("internal_keywords"), list) else [], "nl_status": "无 / 待确认", "reference_model": import_text(product_data.get("reference_model")) or None, "application": import_text(product_data.get("application")) or None, "technical_summary": import_text(product_data.get("technical_summary")) or None, "customizable": import_text(product_data.get("customizable")) if import_text(product_data.get("customizable")) in {"是", "否"} else "待确认", "sample_available": import_text(product_data.get("sample_available")) if import_text(product_data.get("sample_available")) in {"是", "否"} else "待确认", "capacity": import_text(product_data.get("capacity")) or None, "moq": import_text(product_data.get("moq")) or None, "standard_lead_time": import_text(product_data.get("standard_lead_time")) or None, "packaging": import_text(product_data.get("packaging")) or None, "export_capacity": import_text(product_data.get("export_capacity")) or "待确认", "notes": "由 AI 导入；不是正式 NL 牌号。", "import_batch_id": batch_id}
        supplier_product = (await supabase("supplier_products", token, "POST", product_values))[0]
        await add_import_effect(token, batch_id, entity_type="supplier_product", action="created", record_id=supplier_product["id"], before_data=None, after_data=product_values)
    link = None; rfq = None; link_data = payload.get("link") if isinstance(payload.get("link"), dict) else {}; rfq_data = payload.get("rfq") if isinstance(payload.get("rfq"), dict) else {}
    reference = link_data or rfq_data
    if payload.get("intent") in {"link_supplier_to_customer_project", "create_supplier_rfq"}:
        customer_name = import_text(reference.get("customer_company_name")); project_name = import_text(reference.get("project_name"))
        customer_rows = await supabase(f"customers?company_name=ilike.{quote(customer_name, safe='')}&archived_at=is.null&select=id,company_name&limit=2", token)
        if len(customer_rows) != 1: raise HTTPException(422, "supplier import must name one accessible overseas customer")
        project_rows = await supabase(f"projects?customer_id=eq.{customer_rows[0]['id']}&project_name=ilike.{quote(project_name, safe='')}&archived_at=is.null&select=id&limit=2", token)
        if len(project_rows) != 1: raise HTTPException(422, "supplier import must name one accessible customer project")
        link_values = {"customer_id": customer_rows[0]["id"], "project_id": project_rows[0]["id"], "supplier_id": supplier["id"], "supplier_product_id": supplier_product.get("id") if supplier_product else None, "customer_need": import_text(reference.get("customer_need")) or None, "reference_product": import_text(reference.get("reference_product")) or None, "match_status": import_text(reference.get("match_status")) or "待询价", "technical_match_notes": import_text(reference.get("technical_match_notes")) or None, "quote_status": import_text(reference.get("quote_status")) or None, "sample_status": import_text(reference.get("sample_status")) or None, "current_risk": import_text(reference.get("current_risk")) or None, "next_action": import_text(reference.get("next_action")) or None, "next_followup_date": import_date(reference.get("next_followup_date"), "link.next_followup_date"), "import_batch_id": batch_id}
        link = (await supabase("supplier_project_links", token, "POST", link_values))[0]
        await add_import_effect(token, batch_id, entity_type="supplier_project_link", action="created", record_id=link["id"], before_data=None, after_data=link_values)
        if payload.get("intent") == "create_supplier_rfq":
            rfq_values = {"rfq_number": await next_supplier_rfq_number(token), "customer_id": link_values["customer_id"], "project_id": link_values["project_id"], "supplier_id": supplier["id"], "supplier_product_id": supplier_product.get("id") if supplier_product else None, "demand_product": import_text(rfq_data.get("demand_product")) or "待确认", "reference_product": import_text(rfq_data.get("reference_product")) or link_values["reference_product"], "end_application": import_text(rfq_data.get("end_application")) or None, "technical_requirements": import_text(rfq_data.get("technical_requirements")) or None, "sample_quantity": import_text(rfq_data.get("sample_quantity")) or None, "expected_monthly_usage": import_text(rfq_data.get("expected_monthly_usage")) or None, "expected_annual_usage": import_text(rfq_data.get("expected_annual_usage")) or None, "destination_country": import_text(rfq_data.get("destination_country")) or None, "requested_materials": rfq_data.get("requested_materials") if isinstance(rfq_data.get("requested_materials"), list) else [], "status": import_text(rfq_data.get("status")) or "草稿", "created_date": source_date, "sent_date": import_date(rfq_data.get("sent_date"), "rfq.sent_date"), "next_followup_date": import_date(rfq_data.get("next_followup_date"), "rfq.next_followup_date"), "reply_content": import_text(rfq_data.get("reply_content")) or None, "import_batch_id": batch_id}
            rfq = (await supabase("supplier_rfqs", token, "POST", rfq_values))[0]
            await add_import_effect(token, batch_id, entity_type="supplier_rfq", action="created", record_id=rfq["id"], before_data=None, after_data=rfq_values)
    followup = None; followup_data = payload.get("follow_up") if isinstance(payload.get("follow_up"), dict) else {}
    if payload.get("intent") == "upsert_supplier_and_followup" and import_text(followup_data.get("content")):
        followup_values = {"supplier_id": supplier["id"], "date": import_date(followup_data.get("date"), "follow_up.date") or source_date, "channel": import_text(followup_data.get("channel")) or "微信", "content": import_text(followup_data.get("content")), "conclusion": import_text(followup_data.get("conclusion")) or None, "next_action": import_text(followup_data.get("next_action")) or values["next_action"], "next_followup_date": import_date(followup_data.get("next_followup_date"), "follow_up.next_followup_date") or values["next_followup_date"], "status": import_text(followup_data.get("status")) or values["current_status"], "import_batch_id": batch_id}
        followup = (await supabase("supplier_followups", token, "POST", followup_values))[0]
        await add_import_effect(token, batch_id, entity_type="supplier_followup", action="created", record_id=followup["id"], before_data=None, after_data=followup_values)
    await record_timeline_event(token, title=f"AI 导入供应商：{supplier['company_name']}", event_type="crm", source="ai_supplier_import", related_id=supplier["id"], supplier_id=supplier["id"], event_date=source_date)
    await supabase(f"import_batches?id=eq.{batch_id}", token, "PATCH", {"status": "applied", "applied_at": datetime.now().isoformat(), "preview": preview})
    return {"batch_id": batch_id, "supplier": supplier, "supplier_action": supplier_action, "supplier_product": supplier_product, "link": link, "rfq": rfq, "followup": followup}

@app.post("/api/imports/{batch_id}/apply")
async def apply_import(batch_id: str, confirmation: ImportApplyIn, authorization: str | None = Header(default=None)):
    """Apply one approved draft. Every created or changed row gets an import effect log."""
    token = bearer(authorization)
    batch_rows = await supabase(f"import_batches?id=eq.{batch_id}&select=*", token)
    if not batch_rows:
        raise HTTPException(404, "Import batch not found")
    batch = batch_rows[0]
    if batch.get("status") != "draft":
        raise HTTPException(409, "Only a draft import can be applied")
    payload = batch.get("raw_payload") or {}
    if payload.get("intent") in {"upsert_supplier", "upsert_supplier_and_followup", "create_supplier_rfq", "link_supplier_to_customer_project"}:
        return await apply_supplier_import(token, batch_id, payload, confirmation)
    preview = await build_import_preview(token, payload)
    match = preview["customer_match"]
    customer_data = payload.get("customer") or {}
    source = payload.get("source") or {}
    source_date = import_date(source.get("date"), "source.date", required=True) or str(date.today())
    selected_customer_id = confirmation.selected_customer_id or match.get("customer_id")
    if match["kind"] == "internal_forwarder" and not confirmation.selected_customer_id:
        raise HTTPException(422, "内部同事转发邮件必须人工选择真实客户，不能自动创建客户")
    if match["kind"] in ("company_manual_review", "company_ambiguous") and not (confirmation.confirm_company_match or confirmation.selected_customer_id):
        raise HTTPException(422, "公司名称匹配必须人工确认后才能更新客户")
    if match["kind"] == "company_ambiguous" and not confirmation.selected_customer_id:
        raise HTTPException(422, "请从同名客户中选择一个目标客户")

    active_customers = await supabase("customers?select=*&import_reverted=eq.false&archived_at=is.null&limit=500", token)
    if selected_customer_id:
        customer_rows = [row for row in active_customers if row["id"] == selected_customer_id]
        if not customer_rows:
            raise HTTPException(422, "所选客户不存在或已撤销")
        customer = customer_rows[0]
        customer_values = import_nonempty({
            "company_name": import_text(customer_data.get("company_name")), "country": import_text(customer_data.get("country")),
            "contact_person": import_text(customer_data.get("contact_name")), "email": import_text(customer_data.get("email")).lower(),
            "whatsapp": import_text(customer_data.get("wechat_or_whatsapp")), "industry": import_text(customer_data.get("industry")),
            "customer_summary": import_text(customer_data.get("summary")), "customer_background": import_text(customer_data.get("background")),
            "customer_need": import_text(customer_data.get("current_need")), "priority": {"high": "HIGH", "medium": "MEDIUM", "low": "MEDIUM"}.get(import_key(customer_data.get("priority")), import_text(customer_data.get("priority"))),
            "customer_value": customer_data.get("customer_value") if isinstance(customer_data.get("customer_value"), int) else None,
            "customer_stage": import_text(customer_data.get("stage")), "status_label": import_text(customer_data.get("status_text")),
            "customer_tags": customer_data.get("tags") if isinstance(customer_data.get("tags"), list) else None,
            "next_action": [import_text(customer_data.get("next_action"))] if import_text(customer_data.get("next_action")) else None,
            "next_followup_date": import_date(customer_data.get("next_follow_up_date"), "customer.next_follow_up_date"),
            "last_contact_date": source_date,
        })
        # A forwarded internal address may identify context, but it must never
        # overwrite the real customer's CRM email.
        if is_internal_mail_address(import_text(customer_data.get("email") or (payload.get("match") or {}).get("customer_email"))):
            customer_values.pop("email", None)
        before = {key: customer.get(key) for key in customer_values}
        customer = (await supabase(f"customers?id=eq.{customer['id']}", token, "PATCH", customer_values))[0]
        await add_import_effect(token, batch_id, entity_type="customer", action="updated", record_id=customer["id"], before_data=before, after_data=customer_values)
        customer_action = "updated"
    else:
        company_name = import_text(customer_data.get("company_name") or (payload.get("match") or {}).get("company_name"))
        email = import_text(customer_data.get("email") or (payload.get("match") or {}).get("customer_email")).lower()
        if not company_name:
            raise HTTPException(422, "新建客户必须提供 company_name")
        if is_internal_mail_address(email):
            raise HTTPException(422, "内部同事邮箱不能创建为客户")
        customer_values = {
            "company_name": company_name, "country": import_text(customer_data.get("country")) or "待确认",
            "contact_person": import_text(customer_data.get("contact_name")) or "待确认", "email": email or "待确认",
            "whatsapp": import_text(customer_data.get("wechat_or_whatsapp")) or None, "industry": import_text(customer_data.get("industry")) or None,
            "customer_summary": import_text(customer_data.get("summary")) or None, "customer_background": import_text(customer_data.get("background")) or None,
            "customer_need": import_text(customer_data.get("current_need")) or None,
            "priority": {"high": "HIGH", "medium": "MEDIUM", "low": "MEDIUM"}.get(import_key(customer_data.get("priority")), "MEDIUM"),
            "customer_value": customer_data.get("customer_value") if isinstance(customer_data.get("customer_value"), int) else 3,
            "customer_stage": import_text(customer_data.get("stage")) or "New", "status_label": import_text(customer_data.get("status_text")) or None,
            "customer_tags": customer_data.get("tags") if isinstance(customer_data.get("tags"), list) else [],
            "next_action": [import_text(customer_data.get("next_action"))] if import_text(customer_data.get("next_action")) else [],
            "next_followup_date": import_date(customer_data.get("next_follow_up_date"), "customer.next_follow_up_date"),
            "last_contact_date": source_date, "import_batch_id": batch_id,
        }
        customer = (await supabase("customers", token, "POST", customer_values))[0]
        await add_import_effect(token, batch_id, entity_type="customer", action="created", record_id=customer["id"], before_data=None, after_data=customer_values)
        customer_action = "created"

    product_refs = payload.get("product_refs") or []
    product_records: list[dict[str, Any]] = []
    active_products = await supabase("products?select=*&import_reverted=eq.false&archived_at=is.null&limit=500", token)
    for product_data in product_refs:
        code = import_text(product_data.get("code")).upper()
        existing = next((row for row in active_products if import_key(row.get("product_code")) == import_key(code)), None)
        values = import_nonempty({"product_name": import_text(product_data.get("name")) or code, "product_code": code, "category": import_text(product_data.get("category")), "application": import_text(product_data.get("application"))})
        if existing:
            before = {key: existing.get(key) for key in values}
            product = (await supabase(f"products?id=eq.{existing['id']}", token, "PATCH", values))[0]
            await add_import_effect(token, batch_id, entity_type="product", action="updated", record_id=product["id"], before_data=before, after_data=values)
        else:
            product = (await supabase("products", token, "POST", {**values, "notes": "由 AI 导入暂存箱创建", "import_batch_id": batch_id}))[0]
            await add_import_effect(token, batch_id, entity_type="product", action="created", record_id=product["id"], before_data=None, after_data=values)
        product_records.append(product)
        relation_rows = await supabase(f"product_customer_relations?product_id=eq.{product['id']}&customer_id=eq.{customer['id']}&select=*&limit=1", token)
        if not relation_rows:
            relation = (await supabase("product_customer_relations", token, "POST", {"product_id": product["id"], "customer_id": customer["id"], "import_batch_id": batch_id}))[0]
            await add_import_effect(token, batch_id, entity_type="product_customer_relation", action="created", record_id=relation["id"], before_data=None, after_data={"product_id": product["id"], "customer_id": customer["id"]})

    product = product_records[0] if product_records else None
    project_data = payload.get("project") or {}
    project = None
    project_name = import_text(project_data.get("name"))
    if project_name:
        project_rows = await supabase(f"projects?customer_id=eq.{customer['id']}&import_reverted=eq.false&archived_at=is.null&select=*", token)
        existing_project = next((row for row in project_rows if import_key(row.get("project_name")) == import_key(project_name)), None)
        project_values = import_nonempty({"project_name": project_name, "product_id": product.get("id") if product else None, "application": import_text(project_data.get("application")), "stage": import_text(project_data.get("stage")) or import_text(customer_data.get("stage")), "notes": import_text(project_data.get("notes"))})
        if existing_project:
            before = {key: existing_project.get(key) for key in project_values}
            project = (await supabase(f"projects?id=eq.{existing_project['id']}", token, "PATCH", project_values))[0]
            await add_import_effect(token, batch_id, entity_type="project", action="updated", record_id=project["id"], before_data=before, after_data=project_values)
        else:
            project = (await supabase("projects", token, "POST", {**project_values, "customer_id": customer["id"], "import_batch_id": batch_id}))[0]
            await add_import_effect(token, batch_id, entity_type="project", action="created", record_id=project["id"], before_data=None, after_data=project_values)

    followup_data = payload.get("follow_up") or {}
    followup = None
    if import_text(followup_data.get("content")):
        followup_values = {"customer_id": customer["id"], "date": import_date(followup_data.get("date"), "follow_up.date") or source_date, "content": import_text(followup_data.get("content")), "next_action": import_text(followup_data.get("next_action")) or import_text(customer_data.get("next_action")) or "安排下一步行动", "status": import_text(followup_data.get("status")) or "Open", "import_batch_id": batch_id}
        followup = (await supabase("followups", token, "POST", followup_values))[0]
        await add_import_effect(token, batch_id, entity_type="followup", action="created", record_id=followup["id"], before_data=None, after_data=followup_values)

    task_data = payload.get("task") or {}
    task = None
    if task_data.get("create") is True:
        title = import_text(task_data.get("title"))
        if not title:
            raise HTTPException(422, "task.title is required when task.create is true")
        task_values = {"title": title, "description": import_text(task_data.get("description")) or None, "category": import_text(task_data.get("category")) or "外贸", "priority": {"重要": "important", "普通": "normal", "低": "low"}.get(import_text(task_data.get("priority")), "normal"), "status": "Pending", "task_date": import_date(task_data.get("due_date"), "task.due_date") or source_date, "customer_id": customer["id"], "project_id": project.get("id") if project else None, "product_id": product.get("id") if product else None, "import_batch_id": batch_id}
        task = (await supabase("tasks", token, "POST", task_values))[0]
        await add_import_effect(token, batch_id, entity_type="task", action="created", record_id=task["id"], before_data=None, after_data=task_values)

    timeline_values = {"event_date": source_date, "event_time": datetime.now().strftime("%H:%M:%S"), "title": f"AI 导入确认：{customer.get('company_name')} · {import_text(followup_data.get('content')) or import_text(customer_data.get('next_action')) or '更新客户资料'}", "event_type": "crm", "source": "ai_import", "related_id": followup.get("id") if followup else customer["id"], "customer_id": customer["id"], "project_id": project.get("id") if project else None, "product_id": product.get("id") if product else None, "import_batch_id": batch_id}
    timeline = (await supabase("timeline_events", token, "POST", timeline_values))[0]
    await add_import_effect(token, batch_id, entity_type="timeline", action="created", record_id=timeline["id"], before_data=None, after_data=timeline_values)
    await supabase(f"import_batches?id=eq.{batch_id}", token, "PATCH", {"status": "applied", "applied_at": datetime.now().isoformat(), "preview": preview})
    return {"batch_id": batch_id, "customer": customer, "customer_action": customer_action, "project": project, "products": product_records, "followup": followup, "task": task}

@app.post("/api/imports/{batch_id}/revert")
async def revert_import(batch_id: str, authorization: str | None = Header(default=None)):
    """Reverse one batch without hard-deleting any business row."""
    token = bearer(authorization)
    batch_rows = await supabase(f"import_batches?id=eq.{batch_id}&select=*", token)
    if not batch_rows:
        raise HTTPException(404, "Import batch not found")
    if batch_rows[0].get("status") != "applied":
        raise HTTPException(409, "Only an applied import can be reverted")
    effects = await supabase(f"import_effects?import_batch_id=eq.{batch_id}&reverted_at=is.null&select=*&order=created_at.desc", token)
    table_by_entity = {
        "customer": "customers", "project": "projects", "product": "products",
        "product_customer_relation": "product_customer_relations", "followup": "followups",
        "task": "tasks", "timeline": "timeline_events", "supplier": "suppliers",
        "supplier_contact": "supplier_contacts", "supplier_product": "supplier_products",
        "supplier_followup": "supplier_followups", "supplier_project_link": "supplier_project_links",
        "supplier_rfq": "supplier_rfqs",
    }
    now = datetime.now().isoformat()
    for effect in effects:
        table = table_by_entity.get(effect["entity_type"])
        if not table:
            continue
        if effect["action"] == "updated":
            before = effect.get("before_data") or {}
            if before:
                await supabase(f"{table}?id=eq.{effect['record_id']}", token, "PATCH", before)
        else:
            await supabase(f"{table}?id=eq.{effect['record_id']}", token, "PATCH", {"import_reverted": True})
        await supabase(f"import_effects?id=eq.{effect['id']}", token, "PATCH", {"reverted_at": now})
    await supabase(f"import_batches?id=eq.{batch_id}", token, "PATCH", {"status": "reverted", "reverted_at": now})
    return {"batch_id": batch_id, "reverted_effects": len(effects), "message": "已撤销本次导入影响；业务记录未被硬删除。"}

@app.get("/api/quotes")
async def list_quotes(authorization: str | None = Header(default=None)):
    return await supabase("quotes?select=*&archived_at=is.null&order=created_at.desc", bearer(authorization))

@app.post("/api/quotes", status_code=201)
async def create_quote(quote: QuoteIn, authorization: str | None = Header(default=None)):
    return await supabase("quotes", bearer(authorization), "POST", quote.model_dump(exclude_none=True))

# Supplier Center -----------------------------------------------------------
# Supplier products are intentionally separate from public.products.  A formal
# product link is optional and cannot be inferred from a supplier reference.

async def require_supplier(token: str, supplier_id: str) -> dict[str, Any]:
    rows = await supabase(f"suppliers?id=eq.{supplier_id}&archived_at=is.null&import_reverted=eq.false&select=*&limit=1", token)
    if not rows:
        raise HTTPException(404, "Supplier not found")
    return rows[0]

async def require_customer_project(token: str, customer_id: str, project_id: str) -> dict[str, Any]:
    rows = await supabase(f"projects?id=eq.{project_id}&customer_id=eq.{customer_id}&archived_at=is.null&import_reverted=eq.false&select=*&limit=1", token)
    if not rows:
        raise HTTPException(422, "Customer project was not found or is not accessible")
    return rows[0]

@app.get("/api/suppliers")
async def list_suppliers(authorization: str | None = Header(default=None), limit: int = Query(500, le=500)):
    return await supabase(f"suppliers?select=*&archived_at=is.null&import_reverted=eq.false&order=next_followup_date.asc.nullslast,created_at.desc&limit={limit}", bearer(authorization))

@app.post("/api/suppliers", status_code=201)
async def create_supplier(payload: SupplierIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    company = payload.company_name.strip()
    existing = await supabase(f"suppliers?company_name=ilike.{quote(company, safe='')}&archived_at=is.null&import_reverted=eq.false&select=id,company_name,main_email,main_phone&limit=20", token)
    email = import_key(payload.main_email)
    phone = import_key(payload.main_phone)
    duplicate = next((row for row in existing if not email and not phone or (email and import_key(row.get('main_email')) == email) or (phone and import_key(row.get('main_phone')) == phone)), None)
    if duplicate:
        raise HTTPException(409, f"Supplier may already exist: {duplicate['company_name']}. Please open and update the existing supplier instead.")
    rows = await supabase("suppliers", token, "POST", {**payload.model_dump(exclude_none=True), "company_name": company, "updated_at": datetime.now().isoformat()})
    await record_timeline_event(token, title=f"新增供应商：{company}", event_type="crm", source="supplier", related_id=rows[0]["id"], supplier_id=rows[0]["id"])
    return rows[0]

@app.patch("/api/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, payload: SupplierUpdateIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    await require_supplier(token, supplier_id)
    rows = await supabase(f"suppliers?id=eq.{supplier_id}", token, "PATCH", {**payload.model_dump(exclude_none=True), "company_name": payload.company_name.strip(), "updated_at": datetime.now().isoformat()})
    return rows[0]

@app.get("/api/suppliers/{supplier_id}/contacts")
async def list_supplier_contacts(supplier_id: str, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, supplier_id)
    return await supabase(f"supplier_contacts?supplier_id=eq.{supplier_id}&select=*&order=is_primary.desc,created_at.asc", token)

@app.post("/api/supplier-contacts", status_code=201)
async def create_supplier_contact(payload: SupplierContactIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, payload.supplier_id)
    if payload.is_primary:
        await supabase(f"supplier_contacts?supplier_id=eq.{payload.supplier_id}&is_primary=eq.true", token, "PATCH", {"is_primary": False})
    rows = await supabase("supplier_contacts", token, "POST", payload.model_dump(exclude_none=True))
    return rows[0]

@app.get("/api/suppliers/{supplier_id}/products")
async def list_supplier_products(supplier_id: str, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, supplier_id)
    return await supabase(f"supplier_products?supplier_id=eq.{supplier_id}&select=*&order=created_at.desc", token)

@app.post("/api/supplier-products", status_code=201)
async def create_supplier_product(payload: SupplierProductIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, payload.supplier_id)
    # Never allow a formal product id unless the caller explicitly marked it confirmed.
    if payload.nl_product_id and payload.nl_status != "已确认关联":
        raise HTTPException(422, "A formal NL product can be linked only after explicit confirmation")
    if payload.nl_product_id:
        product = await supabase(f"products?id=eq.{payload.nl_product_id}&archived_at=is.null&select=id,product_code&limit=1", token)
        if not product or not str(product[0].get("product_code") or "").upper().startswith("NL-"):
            raise HTTPException(422, "nl_product_id must refer to an existing formal NL product")
    rows = await supabase("supplier_products", token, "POST", payload.model_dump(exclude_none=True))
    return rows[0]

@app.get("/api/suppliers/{supplier_id}/followups")
async def list_supplier_followups(supplier_id: str, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, supplier_id)
    return await supabase(f"supplier_followups?supplier_id=eq.{supplier_id}&select=*&order=date.desc,created_at.desc", token)

@app.post("/api/supplier-followups", status_code=201)
async def create_supplier_followup(payload: SupplierFollowupIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, payload.supplier_id)
    values = payload.model_dump(exclude_none=True, exclude={"create_task"})
    rows = await supabase("supplier_followups", token, "POST", values)
    supplier = (await supabase(f"suppliers?id=eq.{payload.supplier_id}", token, "PATCH", {
        "current_status": payload.status, "last_contact_date": payload.date,
        "next_action": payload.next_action, "next_followup_date": payload.next_followup_date,
        "updated_at": datetime.now().isoformat(),
    }))[0]
    task = None
    if payload.create_task and payload.next_followup_date:
        task_rows = await supabase("tasks", token, "POST", {
            "title": f"跟进供应商：{supplier['company_name']}", "description": payload.next_action or payload.content,
            "category": "外贸", "priority": "normal", "status": "Pending", "task_date": payload.next_followup_date,
            "supplier_id": payload.supplier_id,
        })
        task = task_rows[0]
    await record_timeline_event(token, title=f"供应商跟进：{supplier['company_name']} · {payload.content[:80]}", event_type="crm", source="supplier_followup", related_id=rows[0]["id"], supplier_id=payload.supplier_id, event_date=payload.date)
    return {"followup": rows[0], "supplier": supplier, "task": task}

@app.get("/api/supplier-project-links")
async def list_supplier_project_links(authorization: str | None = Header(default=None), project_id: str | None = None, supplier_id: str | None = None):
    filters = ["select=*", "order=next_followup_date.asc.nullslast,created_at.desc"]
    if project_id: filters.append(f"project_id=eq.{project_id}")
    if supplier_id: filters.append(f"supplier_id=eq.{supplier_id}")
    return await supabase(f"supplier_project_links?{'&'.join(filters)}", bearer(authorization))

@app.post("/api/supplier-project-links", status_code=201)
async def create_supplier_project_link(payload: SupplierProjectLinkIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    await require_supplier(token, payload.supplier_id)
    await require_customer_project(token, payload.customer_id, payload.project_id)
    if payload.supplier_product_id:
        products = await supabase(f"supplier_products?id=eq.{payload.supplier_product_id}&supplier_id=eq.{payload.supplier_id}&select=id&limit=1", token)
        if not products: raise HTTPException(422, "Supplier product does not belong to this supplier")
    existing = await supabase(f"supplier_project_links?project_id=eq.{payload.project_id}&supplier_id=eq.{payload.supplier_id}&supplier_product_id={'eq.' + payload.supplier_product_id if payload.supplier_product_id else 'is.null'}&select=*&limit=1", token)
    values = {**payload.model_dump(exclude_none=True), "updated_at": datetime.now().isoformat()}
    if existing:
        rows = await supabase(f"supplier_project_links?id=eq.{existing[0]['id']}", token, "PATCH", values)
        return rows[0]
    rows = await supabase("supplier_project_links", token, "POST", values)
    await record_timeline_event(token, title="关联供应商至客户项目", event_type="project", source="supplier_link", related_id=rows[0]["id"], customer_id=payload.customer_id, project_id=payload.project_id, supplier_id=payload.supplier_id)
    return rows[0]

async def next_supplier_rfq_number(token: str) -> str:
    prefix = f"SUP-RFQ-{date.today().strftime('%Y%m%d')}-"
    rows = await supabase(f"supplier_rfqs?rfq_number=like.{prefix}*&select=rfq_number&limit=500", token)
    numbers = [int(str(row.get("rfq_number") or "").rsplit("-", 1)[-1]) for row in rows if str(row.get("rfq_number") or "").rsplit("-", 1)[-1].isdigit()]
    return f"{prefix}{max(numbers, default=0) + 1:03d}"

@app.get("/api/supplier-rfqs")
async def list_supplier_rfqs(authorization: str | None = Header(default=None), supplier_id: str | None = None, project_id: str | None = None):
    filters = ["select=*", "order=created_date.desc,created_at.desc"]
    if supplier_id: filters.append(f"supplier_id=eq.{supplier_id}")
    if project_id: filters.append(f"project_id=eq.{project_id}")
    return await supabase(f"supplier_rfqs?{'&'.join(filters)}", bearer(authorization))

@app.get("/api/supplier-insights")
async def supplier_insights(authorization: str | None = Header(default=None)):
    """Small aggregated facts for the supplier list; every source is still RLS-scoped."""
    token = bearer(authorization)
    documents = await supabase("supplier_documents?select=supplier_id,document_type&limit=5000", token)
    products = await supabase("supplier_products?select=supplier_id,sample_available&limit=5000", token)
    rfqs = await supabase("supplier_rfqs?select=supplier_id,status&limit=5000", token)
    links = await supabase("supplier_project_links?select=supplier_id&limit=5000", token)
    by_supplier: dict[str, dict[str, Any]] = {}
    def insight(supplier_id: str) -> dict[str, Any]:
        return by_supplier.setdefault(supplier_id, {"supplier_id": supplier_id, "document_count": 0, "tds_count": 0, "sample_available": False, "effective_quote_count": 0, "project_link_count": 0})
    for row in documents:
        item = insight(row["supplier_id"]); item["document_count"] += 1
        if row.get("document_type") == "TDS": item["tds_count"] += 1
        if row.get("document_type") == "报价单": item["effective_quote_count"] += 1
    for row in products:
        if row.get("sample_available") == "是": insight(row["supplier_id"])["sample_available"] = True
    for row in rfqs:
        if row.get("status") in {"供应商已回复", "技术评估"}: insight(row["supplier_id"])["effective_quote_count"] += 1
    for row in links: insight(row["supplier_id"])["project_link_count"] += 1
    return list(by_supplier.values())

@app.post("/api/supplier-rfqs", status_code=201)
async def create_supplier_rfq(payload: SupplierRfqIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    await require_supplier(token, payload.supplier_id)
    await require_customer_project(token, payload.customer_id, payload.project_id)
    if payload.supplier_product_id:
        products = await supabase(f"supplier_products?id=eq.{payload.supplier_product_id}&supplier_id=eq.{payload.supplier_id}&select=id&limit=1", token)
        if not products: raise HTTPException(422, "Supplier product does not belong to this supplier")
    values = payload.model_dump(exclude_none=True)
    values["rfq_number"] = await next_supplier_rfq_number(token)
    values["created_date"] = str(date.today())
    values["updated_at"] = datetime.now().isoformat()
    rows = await supabase("supplier_rfqs", token, "POST", values)
    await record_timeline_event(token, title=f"创建供应商 RFQ：{values['rfq_number']}", event_type="project", source="supplier_rfq", related_id=rows[0]["id"], customer_id=payload.customer_id, project_id=payload.project_id, supplier_id=payload.supplier_id, supplier_rfq_id=rows[0]["id"])
    return rows[0]

async def supplier_storage_upload(path: str, content: bytes, mime_type: str | None) -> None:
    cfg = settings()
    headers = {"apikey": cfg.supabase_service_role_key, "Authorization": f"Bearer {cfg.supabase_service_role_key}", "Content-Type": mime_type or "application/octet-stream", "x-upsert": "false"}
    async with httpx.AsyncClient(timeout=45) as client:
        response = await client.post(f"{cfg.supabase_url}/storage/v1/object/supplier-documents/{quote(path, safe='/')}", headers=headers, content=content)
    if response.status_code >= 400: raise HTTPException(502, "Supplier document upload failed")

async def supplier_storage_signed_url(path: str) -> str:
    cfg = settings()
    headers = {"apikey": cfg.supabase_service_role_key, "Authorization": f"Bearer {cfg.supabase_service_role_key}", "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=20) as client:
        response = await client.post(f"{cfg.supabase_url}/storage/v1/object/sign/supplier-documents/{quote(path, safe='/')}", headers=headers, json={"expiresIn": 3600})
    if response.status_code >= 400: raise HTTPException(502, "Supplier document preview could not be created")
    signed_path = response.json().get("signedURL")
    if not signed_path: raise HTTPException(502, "Supplier document preview could not be created")
    return f"{cfg.supabase_url}/storage/v1{signed_path}"

@app.get("/api/suppliers/{supplier_id}/documents")
async def list_supplier_documents(supplier_id: str, authorization: str | None = Header(default=None)):
    token = bearer(authorization); await require_supplier(token, supplier_id)
    return await supabase(f"supplier_documents?supplier_id=eq.{supplier_id}&select=*&order=uploaded_at.desc", token)

@app.post("/api/supplier-documents", status_code=201)
async def upload_supplier_document(
    supplier_id: str = Form(...), document_type: str = Form(...), file: UploadFile = File(...),
    supplier_product_id: str | None = Form(default=None), project_id: str | None = Form(default=None),
    rfq_id: str | None = Form(default=None), source: str = Form(default="手动上传"), internal_notes: str | None = Form(default=None),
    authorization: str | None = Header(default=None),
):
    token = bearer(authorization); await require_supplier(token, supplier_id)
    allowed = {"TDS", "SDS", "COA", "报价单", "产品图片", "认证文件", "邮件附件", "其他资料"}
    if document_type not in allowed: raise HTTPException(422, "Unsupported supplier document type")
    content = await file.read()
    if not content or len(content) > 25 * 1024 * 1024: raise HTTPException(422, "Document must be between 1 byte and 25 MB")
    safe_name = "".join(character if character.isalnum() or character in ".-_" else "_" for character in (file.filename or "document"))
    storage_path = f"{supplier_id}/{datetime.now().strftime('%Y%m%d%H%M%S%f')}-{safe_name}"
    await supplier_storage_upload(storage_path, content, file.content_type)
    rows = await supabase("supplier_documents", token, "POST", {"supplier_id": supplier_id, "supplier_product_id": supplier_product_id, "project_id": project_id, "rfq_id": rfq_id, "document_type": document_type, "file_name": file.filename or safe_name, "storage_path": storage_path, "mime_type": file.content_type, "file_size": len(content), "source": source, "internal_notes": internal_notes})
    return rows[0]

@app.get("/api/supplier-documents/{document_id}/preview")
async def preview_supplier_document(document_id: str, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    rows = await supabase(f"supplier_documents?id=eq.{document_id}&select=*&limit=1", token)
    if not rows: raise HTTPException(404, "Supplier document not found")
    return {"url": await supplier_storage_signed_url(rows[0]["storage_path"]), "file_name": rows[0]["file_name"], "mime_type": rows[0].get("mime_type")}

@app.get("/api/emails")
async def list_emails(
    authorization: str | None = Header(default=None), limit: int = Query(500, le=500),
    category: Literal["customer_inquiry", "technical", "quotation", "sample", "payment", "other"] | None = None,
    status: Literal["unread", "new_lead", "linked", "followup_created", "completed"] | None = None,
    unlinked: bool = False,
):
    filters = ["select=*", "order=received_at.desc", f"limit={limit}"]
    if category:
        filters.append(f"category=eq.{category}")
    if status:
        filters.append(f"status=eq.{status}")
    if unlinked:
        filters.append("customer_id=is.null")
    rows = await supabase(f"emails?{'&'.join(filters)}", bearer(authorization))
    return [{**row, "is_internal_sender": is_internal_mail_address(row.get("sender"))} for row in rows]

@app.get("/api/emails/unlinked")
async def list_unlinked_emails(authorization: str | None = Header(default=None), limit: int = Query(200, le=200)):
    rows = await supabase(f"emails?customer_id=is.null&select=*&order=received_at.desc&limit={limit}", bearer(authorization))
    return [{**row, "is_internal_sender": is_internal_mail_address(row.get("sender"))} for row in rows]

@app.get("/api/emails/{email_id}")
async def get_email(email_id: str, authorization: str | None = Header(default=None)):
    rows = await supabase(f"emails?id=eq.{email_id}&select=*", bearer(authorization))
    if not rows:
        raise HTTPException(404, "Email not found")
    return {**rows[0], "is_internal_sender": is_internal_mail_address(rows[0].get("sender"))}

@app.patch("/api/emails/{email_id}")
async def update_email_status(email_id: str, payload: EmailStatusIn, authorization: str | None = Header(default=None)):
    rows = await supabase(f"emails?id=eq.{email_id}", bearer(authorization), "PATCH", payload.model_dump())
    if not rows:
        raise HTTPException(404, "Email not found")
    return {**rows[0], "is_internal_sender": is_internal_mail_address(rows[0].get("sender"))}

@app.post("/api/emails/{email_id}/link")
async def link_email_to_customer(email_id: str, payload: EmailLinkIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    rows = await supabase(f"emails?id=eq.{email_id}&select=*", token)
    if not rows:
        raise HTTPException(404, "Email not found")
    customer_rows = await supabase(f"customers?id=eq.{payload.customer_id}&archived_at=is.null&select=id,contact_person", token)
    if not customer_rows:
        raise HTTPException(404, "Customer not found")
    projects = await supabase(f"projects?customer_id=eq.{payload.customer_id}&archived_at=is.null&select=id,product_id&order=created_at.desc&limit=1", token)
    project = projects[0] if projects else {}
    email = rows[0]
    if not is_internal_mail_address(email.get("sender")):
        mapping_rows = await supabase(f"customer_email_mappings?email_address=eq.{quote(email['sender'], safe='')}&select=id", token)
        mapping_payload = {"customer_id": payload.customer_id, "email_address": email["sender"], "contact_name": payload.contact_name or email.get("sender_name") or customer_rows[0].get("contact_person")}
        if mapping_rows:
            await supabase(f"customer_email_mappings?id=eq.{mapping_rows[0]['id']}", token, "PATCH", mapping_payload)
        else:
            await supabase("customer_email_mappings", token, "POST", mapping_payload)
    updated = await supabase(f"emails?id=eq.{email_id}", token, "PATCH", {
        "customer_id": payload.customer_id,
        "project_id": project.get("id"),
        "product_id": project.get("product_id"),
        "status": "linked",
    })
    return {**updated[0], "is_internal_sender": is_internal_mail_address(updated[0].get("sender"))}

@app.post("/api/emails/{email_id}/customers", status_code=201)
async def create_customer_from_email(email_id: str, payload: EmailCustomerCreateIn, authorization: str | None = Header(default=None)):
    """Create a customer owned by the signed-in member and link the reviewed email."""
    token = bearer(authorization)
    email_rows = await supabase(f"emails?id=eq.{email_id}&select=*", token)
    if not email_rows:
        raise HTTPException(404, "Email not found")
    email = email_rows[0]
    if is_internal_mail_address(email.get("sender")):
        raise HTTPException(422, "Internal colleague mail cannot create a customer")

    address = payload.email.strip().lower()
    existing = await supabase(
        f"customers?email=eq.{quote(address, safe='')}&archived_at=is.null&import_reverted=eq.false&select=id,company_name&limit=1",
        token,
    )
    if existing:
        raise HTTPException(409, f"Customer email already exists: {existing[0]['company_name']}. Link the email to that customer instead.")

    company_name = payload.company_name.strip()
    existing_company = await supabase(
        f"customers?company_name=ilike.{quote(company_name, safe='')}&archived_at=is.null&import_reverted=eq.false&select=id,company_name&limit=1",
        token,
    )
    if existing_company:
        raise HTTPException(409, f"Customer company already exists: {existing_company[0]['company_name']}. Link the email to that customer instead.")

    customer = (await supabase("customers", token, "POST", payload.model_dump(exclude_none=True)))[0]
    await supabase("customer_email_mappings", token, "POST", {
        "customer_id": customer["id"], "email_address": email.get("sender") or address,
        "contact_name": email.get("sender_name") or payload.contact_person,
    })
    updated = (await supabase(f"emails?id=eq.{email_id}", token, "PATCH", {
        "customer_id": customer["id"], "status": "linked",
    }))[0]
    await record_timeline_event(
        token, title=f"从邮件创建客户：{customer['company_name']}", event_type="crm",
        source="mail_new_lead", related_id=email_id, customer_id=customer["id"],
        event_date=str(email.get("received_at") or date.today())[:10],
    )
    return {"customer": customer, "email": {**updated, "is_internal_sender": is_internal_mail_address(updated.get("sender"))}}

@app.post("/api/emails/{email_id}/followups", status_code=201)
async def create_followup_from_email(email_id: str, payload: EmailFollowupIn, authorization: str | None = Header(default=None)):
    token = bearer(authorization)
    rows = await supabase(f"emails?id=eq.{email_id}&select=*", token)
    if not rows:
        raise HTTPException(404, "Email not found")
    email = rows[0]
    if not email.get("customer_id"):
        raise HTTPException(422, "This email is not linked to a CRM customer")
    record = await supabase("followups", token, "POST", {
        "customer_id": email["customer_id"],
        "email_id": email_id,
        "date": str(email.get("received_at") or "")[:10] or None,
        "content": payload.content or f"邮件：{email.get('subject', '(无主题)')}\n{email.get('content_preview') or ''}",
        "next_action": payload.next_action or "阅读邮件并确认下一步行动",
        "status": "Open",
    })
    customer_rows = await supabase(f"customers?id=eq.{email['customer_id']}&archived_at=is.null&select=next_followup_date", token)
    await supabase("email_actions", token, "POST", {
        "email_id": email_id,
        "customer_id": email["customer_id"],
        "action": "Create follow-up",
        "next_action": payload.next_action or "阅读邮件并确认下一步行动",
        "deadline": customer_rows[0].get("next_followup_date") if customer_rows else None,
        "status": "Pending",
    })
    # Turning an email into a CRM follow-up also creates a visible Daily Focus task.
    # The migration is optional during rollout, so an unavailable tasks table must
    # never prevent the user from creating the original CRM follow-up.
    try:
        task_rows = await supabase("tasks", token, "POST", {
            "title": f"回复邮件：{email.get('subject', '(无主题)')[:120]}",
            "description": payload.content or email.get("content_preview") or "由邮件中心创建的跟进任务",
            "category": "外贸", "priority": "normal", "status": "Pending",
            "task_date": customer_rows[0].get("next_followup_date") if customer_rows and customer_rows[0].get("next_followup_date") else str(date.today()),
            "customer_id": email["customer_id"], "project_id": email.get("project_id"), "product_id": email.get("product_id"),
        })
        task = task_rows[0]
        await record_timeline_event(token, title=f"邮件转任务：{task['title']}", event_type="email", source="mail_followup", related_id=task["id"], customer_id=task.get("customer_id"), project_id=task.get("project_id"), product_id=task.get("product_id"), event_date=task["task_date"])
    except HTTPException:
        pass
    await supabase(f"emails?id=eq.{email_id}", token, "PATCH", {"status": "followup_created"})
    return record[0]

@app.post("/api/emails/{email_id}/update-crm")
async def update_crm_from_email(email_id: str, payload: EmailCrmUpdateIn, authorization: str | None = Header(default=None)):
    """Turn one reviewed email into a complete CRM update without duplicate entry."""
    token = bearer(authorization)
    email_rows = await supabase(f"emails?id=eq.{email_id}&select=*", token)
    if not email_rows:
        raise HTTPException(404, "Email not found")
    email = email_rows[0]
    customer_rows = await supabase(f"customers?id=eq.{payload.customer_id}&archived_at=is.null&select=*", token)
    if not customer_rows:
        raise HTTPException(404, "Customer not found")
    customer = customer_rows[0]
    project = None
    if payload.project_id:
        project_rows = await supabase(f"projects?id=eq.{payload.project_id}&customer_id=eq.{payload.customer_id}&archived_at=is.null&select=*", token)
        if not project_rows:
            raise HTTPException(422, "Selected project does not belong to this customer")
        project = (await supabase(f"projects?id=eq.{payload.project_id}&archived_at=is.null", token, "PATCH", {
            "product_id": payload.product_id,
            "stage": payload.customer_stage,
        }))[0]
    if payload.product_id:
        product_rows = await supabase(f"products?id=eq.{payload.product_id}&archived_at=is.null&select=product_code", token)
        if not product_rows:
            raise HTTPException(404, "Product not found")
        product_code = product_rows[0]["product_code"]
    else:
        product_code = customer.get("product_interest")
    updated_customer = (await supabase(f"customers?id=eq.{payload.customer_id}&archived_at=is.null", token, "PATCH", {
        "customer_stage": payload.customer_stage,
        "next_followup_date": payload.followup_date,
        "next_action": [payload.next_action],
        "status_label": payload.next_action,
        "product_interest": product_code,
    }))[0]
    if not is_internal_mail_address(email.get("sender")):
        mapping_rows = await supabase(f"customer_email_mappings?email_address=eq.{quote(email['sender'], safe='')}&select=id", token)
        mapping_payload = {"customer_id": payload.customer_id, "email_address": email["sender"], "contact_name": email.get("sender_name") or customer.get("contact_person")}
        if mapping_rows:
            await supabase(f"customer_email_mappings?id=eq.{mapping_rows[0]['id']}", token, "PATCH", mapping_payload)
        else:
            await supabase("customer_email_mappings", token, "POST", mapping_payload)
    email_updated = (await supabase(f"emails?id=eq.{email_id}", token, "PATCH", {
        "customer_id": payload.customer_id, "project_id": payload.project_id,
        "product_id": payload.product_id, "status": "followup_created",
    }))[0]
    followup = (await supabase("followups", token, "POST", {
        "customer_id": payload.customer_id, "email_id": email_id, "date": payload.followup_date,
        "content": payload.notes, "next_action": payload.next_action, "status": "Open",
    }))[0]
    await supabase("email_actions", token, "POST", {
        "email_id": email_id, "customer_id": payload.customer_id, "action": "Update CRM",
        "next_action": payload.next_action, "deadline": payload.followup_date, "status": "Pending",
    })
    await record_timeline_event(token, title=f"邮件更新 CRM：{email.get('subject', '(无主题)')[:120]}", event_type="crm", source="mail_crm_update", related_id=followup["id"], customer_id=payload.customer_id, project_id=payload.project_id, product_id=payload.product_id, event_date=payload.followup_date)
    task = None
    if payload.create_task:
        try:
            task = (await supabase("tasks", token, "POST", {
                "title": payload.next_action, "description": payload.notes, "category": "外贸", "priority": "important", "status": "Pending",
                "task_date": payload.task_date or payload.followup_date, "customer_id": payload.customer_id,
                "project_id": payload.project_id, "product_id": payload.product_id,
            }))[0]
            await record_timeline_event(token, title=f"邮件转任务：{task['title']}", event_type="task", source="mail_crm_update", related_id=task["id"], customer_id=payload.customer_id, project_id=payload.project_id, product_id=payload.product_id, event_date=task["task_date"])
        except HTTPException:
            # Daily Focus tables may be introduced later; the CRM update remains valid.
            task = None
    return {"email": {**email_updated, "is_internal_sender": is_internal_mail_address(email_updated.get("sender"))}, "customer": updated_customer, "project": project, "followup": followup, "task": task}

@app.get("/api/email-sync")
async def get_email_sync(authorization: str | None = Header(default=None)):
    rows = await supabase("email_sync?select=*&limit=1", bearer(authorization))
    return rows[0] if rows else {"status": "Not configured", "total_synced": 0, "last_sync_time": None}

@app.get("/api/mailbox")
async def get_current_mailbox(authorization: str | None = Header(default=None)):
    """Return only the signed-in member's safe mailbox label, never IMAP secrets."""
    token = bearer(authorization)
    try:
        rows = await supabase("mailbox_accounts?select=id,mailbox_key,label,email_address,is_active&limit=1", token)
    except HTTPException:
        # Keeps the existing single-mailbox release usable until the V1.8 SQL migration is run.
        return {"configured": False, "label": "当前邮件中心", "email_address": None, "is_active": False}
    if not rows:
        return {"configured": False, "label": "当前邮件中心", "email_address": None, "is_active": False}
    return {**rows[0], "configured": True}
